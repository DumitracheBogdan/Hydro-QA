"""
Build the Hydrocert bidirectional-parity EVIDENCE workbook.

Consumes the parity run's `summary.json` plus the per-flow screenshots and
produces a two-sheet Excel report that matches the existing Hydro-QA report
visual language (PALETTE, card() metric tiles, navy title banner, PASS/FAIL/SKIP
status fills, scaled embedded screenshots) used by
`scripts/generate_regression_excel_dashboard.py`,
`scripts/generate_mobile_regression_excel.py` and
`scripts/generate_detector_excel.py`.

Usage:
    python scripts/build_parity_evidence_excel.py <summary.json> <screenshots_dir> <out.xlsx>

summary.json shape:
    {runId, visitRef, total, passed, failed, gateFailed,
     checks:[{id, direction, status PASS/FAIL/SKIP, details, fields?}]}

Screenshots in <screenshots_dir>:
    mobile = {flow}-before.png / {flow}-after.png
    web    = {check}-web-set.png / {check}-web-verify.png   (added by F2; may be absent)

  where {flow} is the full flow-file prefix, e.g.
    p01a_web2mobile_description-after.png

  The check-id <-> flow map below is the source of truth. Check 2c
  (inspection actions) is API-only (known gap F-01) and has NO screenshot.

Sheets:
  1) "Summary"  — title banner + 4 metric cards (Total / Passed / Failed / Gate)
                  + a table: Check | Direction | Status | Connection (API) |
                  Web evidence | Mobile evidence | Detail-link (intra-workbook
                  hyperlink to the matching Evidence row).
  2) "Evidence" — one row per check (same order as Summary) with:
                  Check | Description | Steps to reproduce | Expected | Actual |
                  Connection check (API GET) | Status |
                  Web screenshot (EMBEDDED) | Mobile screenshot (EMBEDDED).
                  Both the web ({check}-web-*.png) and mobile ({flow}-after.png)
                  images are embedded side by side (scaled ~260px). A missing
                  web image renders the text "web screenshot pending (F2)".

The script never crashes on missing screenshots; absent images degrade to
text placeholders. Console output stays ASCII (Windows cp1252 safe).
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Shared visual language (identical to the sibling report generators)
# ---------------------------------------------------------------------------
PALETTE = {
    'navy': '0F172A',
    'slate': '334155',
    'muted': '64748B',
    'border': 'CBD5E1',
    'pass': '0F766E',
    'pass_bg': 'CCFBF1',
    'fail': 'B91C1C',
    'fail_bg': 'FEE2E2',
    'skip': '92400E',
    'skip_bg': 'FEF3C7',
}

# Check-id -> flow-file prefix (used to find {flow}-after.png mobile screenshots).
# Mirrors the map in the task spec / docs/PARITY-FACTS.md "Full-coverage extension".
# 2c is API-only (gap F-01) -> no screenshot.
CHECK_TO_FLOW = {
    '2a-description': 'p01a_web2mobile_description',
    '2b-visit-actions': 'p01b_web2mobile_visit_actions',
    '2d-visit-text': 'p01d_web2mobile_visit_text',
    '2g-item-detail': 'p01e_web2mobile_item_detail',
    '2h-samples': 'p01f_web2mobile_samples',
    '2c-inspection-actions': None,   # API-only, gap F-01, no shot
    '3a-signature': 'p02_mobile2web_signature',
    '3b-visit-info': 'p03_mobile2web_visit_info',
    '3c-risk': 'p04_mobile2web_risk_assessment',
    '3d-visit-text': 'p05_mobile2web_visit_text',
    '3e-site-induction': 'p03b_mobile2web_site_induction',
    '4a-inspection-notes': 'p06_web2mobile_inspection_notes',
    '4c-item-reference': 'p07_web2mobile_item_reference',
    '4d-item-location': 'p08_web2mobile_item_location',
    '4b-booking-info': 'p09_web2mobile_booking_info',
    '2i-add-inspection': 'p10_web2mobile_add_inspection',
    '2j-visit-status': 'p11_web2mobile_visit_status',
    '4e-mobile-action': 'p12_mobile2web_add_action',
    '2k-sample-note': 'p13_web2mobile_sample_note',
    '2l-engineers': 'p14_web2mobile_engineers',
    '4f-ra-dropdowns': 'p15_web2mobile_ra_dropdowns',
}

# Short check token (e.g. "2a") used to find F2 web screenshots {check}-web-set.png.
def _short(check_id: str) -> str:
    return str(check_id).split('-', 1)[0]


# Per-check evidence MAP: human description + concrete reproduction steps +
# expected + the API connection (GET) check. Authored from docs/PARITY-FACTS.md
# and docs/PARITY-CONTRACT.md so each entry is accurate to what the check does.
EVIDENCE_MAP = {
    '2a-description': {
        'description': 'Web->mobile: a visit Description (visit.notes) entered on the '
                       'webapp must appear in the mobile read-only "Description" card.',
        'steps': [
            'API: PATCH /visits/{id} notes = "PARITY-<runId> description" (or type+Save the Description card on the webapp).',
            'Webapp: open the visit details page; confirm the Description card shows the notes text.',
            'Mobile: open the visit, read the read-only "Description" card.',
            'Assert the "PARITY-<runId> description" text is visible on mobile.',
        ],
        'expected': 'Mobile "Description" card shows the same notes text set on the webapp.',
        'connection': 'GET /visits/{id} -> .notes equals "PARITY-<runId> description".',
    },
    '2b-visit-actions': {
        'description': 'Web->mobile: 3 visit-level actions (High/Medium/Low) created on '
                       'the webapp must render on the mobile visit-detail Actions card.',
        'steps': [
            'Webapp: open the visit Actions card -> "New Action"; create 3 actions named "PARITY-<runId> Hi/Med/Lo" with priorities high/medium/low.',
            'Mobile: open the visit, expand the Actions card.',
            'Assert all 3 action rows (Hi, Med, Lo) are visible.',
        ],
        'expected': 'All 3 visit actions (name + priority) render on the mobile Actions card.',
        'connection': 'GET /actions?visitId={id} -> array of 3 (compare name + priority).',
    },
    '2d-visit-text': {
        'description': 'Web->mobile: 3 visit-text fields (waterSystemDescription, workDetails, '
                       'samplingDetails) set via API must show in the mobile "Visit Details" card.',
        'steps': [
            'API: PATCH /visits/{id} waterSystemDescription/workDetails/samplingDetails = "PARITY-<runId> wsd-web / wd-web / sd-web" (CreateVisitDto rejects these; UpdateVisitDto accepts).',
            'Webapp: open visit details; expand the "Visit Details" card; confirm the 3 values render.',
            'Mobile: open the visit, expand the "Visit Details" card (tapOn text:"Visit Details", below:"Description").',
            'Assert Description & Reference / Work Details / Water Sampling Details show the 3 values.',
        ],
        'expected': 'Mobile "Visit Details" card shows all 3 web-set text fields.',
        'connection': 'GET /visits/{id} -> .waterSystemDescription / .workDetails / .samplingDetails.',
    },
    '2g-item-detail': {
        'description': 'Web->mobile: inspection.itemDetail PATCHed via API must render '
                       'read-only on the mobile inspection LocationCard.',
        'steps': [
            'API: PATCH /inspections/{id} itemDetail = "PARITY-<runId> item-detail".',
            'Webapp: open the Inspections tab -> inspection; confirm the rendered itemDetail value.',
            'Mobile: open the inspection; read the LocationCard.',
            'Assert "PARITY-<runId> item-detail" is visible on the LocationCard.',
        ],
        'expected': 'Mobile LocationCard renders the API-set itemDetail value.',
        'connection': 'GET /inspections/{id} -> .itemDetail.',
    },
    '2h-samples': {
        'description': 'Web->mobile: every base water-sample type (16) added to the inspection via the '
                       'web API must propagate to laboratorySamples and render in the mobile Water Sampling section. '
                       'Per-sample verification. NEVER submits to Normec/ALS.',
        'steps': [
            'API: GET /sample-types (16 types); PATCH /inspections/{id} {samples:[{sampleTypeId,quantity:1}] x16} (additive — safe, local/BE only).',
            'Connection: GET /inspections/{id} -> .laboratorySamples; assert all 16 sampleTypeIds present (checkSamples).',
            'Webapp: open the inspection -> Lab Results tab; screenshot the listed samples.',
            'Mobile: open the inspection -> Water Sampling section; screenshot the synced samples.',
            'GUARDRAIL: never tap "Submit Samples" / POST /laboratory-samples/submit-batch.',
        ],
        'expected': 'All 16 sample types present in laboratorySamples and shown on both webapp + mobile.',
        'connection': 'GET /inspections/{id} -> .laboratorySamples[].sampleTypeId (16/16).',
    },
    '2c-inspection-actions': {
        'description': 'Web->mobile (API-only / KNOWN GAP F-01): 3 inspection-level actions '
                       'created via API. They are confirmed present in the backend but do NOT '
                       'render on the mobile inspection Actions card (TankInspectionScreen.kt:727), '
                       'so this check is verified by API only -- no mobile/web screenshot.',
        'steps': [
            'API: POST /actions {inspectionId, siteId, name, priority} x3 (Hi/Med/Lo).',
            'API: GET /actions?inspectionId={id} and compare name + priority of all 3.',
            'NOTE: mobile render is a known gap (F-01) -- the inspection Actions card stays empty for API-created actions, so no UI assertion / screenshot is taken.',
        ],
        'expected': 'API returns all 3 inspection actions (name + priority). Mobile render is a known gap (F-01); no UI proof.',
        'connection': 'GET /actions?inspectionId={id} -> array of 3 (compare name + priority).',
    },
    '3a-signature': {
        'description': 'Mobile->web: a client signature + signer name captured on mobile '
                       'must render on the webapp "Client Signature" card.',
        'steps': [
            'Mobile: open the visit signature pad; type signer name "PARITY-<runId> Client" and draw/Save the signature.',
            'Webapp: open visit details; confirm the "Client Signature" card shows the image + name.',
            'Assert signatureName matches and the signature image is present (non-null).',
        ],
        'expected': 'Webapp "Client Signature" card shows the mobile-entered name + signature image.',
        'connection': 'GET /visits/{id} -> .signatureName (= entered name) + .signature (non-null).',
    },
    '3b-visit-info': {
        'description': 'Mobile->web: Visit Information form fields (Assisting 1/2/3, '
                       'Works being carried out) typed on mobile must render on the webapp.',
        'steps': [
            'Mobile: open the inspection Visit Information form; type the 4 fields (Assisting 1/2/3, Works being carried out) = "PARITY-<runId> ..." and Save.',
            'Webapp: open the Inspections tab -> inspection; confirm the Visit Information values.',
            'Assert all 4 field values match.',
        ],
        'expected': 'Webapp Visit Information form shows all 4 mobile-typed field values.',
        'connection': 'GET /inspections/{id} -> inspectionForms[formName="Visit Information"].formFields[].formField.fieldName -> .value.',
    },
    '3c-risk': {
        'description': 'Mobile->web: Risk Assessment free-text "- Comments" field typed on '
                       'mobile must render on the webapp (CI automates 1 of the 18 RA fields).',
        'steps': [
            'Mobile: open the inspection Risk Assessment form; scroll to "Accessing Area/Lone Working- Comments"; type "PARITY-<runId> ..." and scroll DOWN to Save.',
            'Webapp: open the Inspections tab -> inspection; confirm the Risk Assessment "- Comments" value.',
            'Assert the field value matches. (CI geometry caps automation at 1 RA field; 18 run locally.)',
        ],
        'expected': 'Webapp Risk Assessment "Accessing Area/Lone Working- Comments" shows the mobile-typed value.',
        'connection': 'GET /inspections/{id} -> inspectionForms[formName="Risk Assessment"].formFields[].formField.fieldName -> .value.',
    },
    '3d-visit-text': {
        'description': 'Mobile->web: 3 visit-text fields (waterSystemDescription, workDetails, '
                       'samplingDetails) typed on mobile must render on the webapp "Visit Details" card '
                       '(bidirectional counterpart of 2d, distinct values).',
        'steps': [
            'Mobile: expand the "Visit Details" card; type Description & Reference / Work Details / Water Sampling Details = "PARITY-<runId> watersys / workdetails / sampling"; visit-level Save.',
            'Webapp: open visit details; expand the "Visit Details" card; confirm the 3 rendered values.',
            'Assert all 3 field values match.',
        ],
        'expected': 'Webapp "Visit Details" card shows all 3 mobile-typed text fields.',
        'connection': 'GET /visits/{id} -> .waterSystemDescription / .workDetails / .samplingDetails.',
    },
    '3e-site-induction': {
        'description': 'Mobile->web: the "Site Induction required & Completed" dropdown set on '
                       'mobile must render on the webapp Visit Information form.',
        'steps': [
            'Mobile: open the Visit Information form; in the "Site Induction required & Completed" dropdown select "Yes - Induction completed"; Save.',
            'Webapp: open the Inspections tab -> inspection; confirm the Site Induction value.',
            'Assert the value equals "Yes - Induction completed".',
        ],
        'expected': 'Webapp shows "Site Induction required & Completed" = "Yes - Induction completed".',
        'connection': 'GET /inspections/{id} -> inspectionForms[formName="Visit Information"].formFields[fieldName="Site Induction required & Completed"].value.',
    },
    '4a-inspection-notes': {
        'description': 'Web->mobile (API-set): inspection.notes PATCHed via the web API must read '
                       'back on GET /inspections and render in the mobile inspection Notes card. '
                       'Additive — never touches the visit title/ref/dates/status.',
        'steps': [
            'API: PATCH /inspections/{id} notes = "PARITY-<runId> insp-notes".',
            'Connection: GET /inspections/{id} -> .notes equals the set value (checkScalarField).',
            'Webapp: open the Inspections tab -> inspection; confirm the Notes section value.',
            'Mobile: open the inspection; read the Notes card (p06, photo-only).',
        ],
        'expected': 'GET /inspections.notes equals the API-set value; mobile Notes card shows it.',
        'connection': 'GET /inspections/{id} -> .notes equals "PARITY-<runId> insp-notes".',
    },
    '4b-booking-info': {
        'description': 'Web->mobile (API-set): site.accessInfo (booking/access info) PATCHed via '
                       'PATCH /sites/{siteId} must read back on GET /sites and render on the mobile '
                       'visit booking/access info. GUARDRAIL: edits the SITE record shared by all '
                       'visits at that site (fine on the dev test site, not visit-scoped).',
        'steps': [
            'API: resolve the visit siteId; PATCH /sites/{siteId} accessInfo = "PARITY-<runId> booking".',
            'Connection: GET /sites/{siteId} -> .accessInfo equals the set value (checkScalarField).',
            'Webapp: open the visit details; confirm the Booking Info card value.',
            'Mobile: open the visit; read the booking/access info (p09, photo-only).',
        ],
        'expected': 'GET /sites.accessInfo equals the API-set value; webapp Booking Info card shows it.',
        'connection': 'GET /sites/{siteId} -> .accessInfo equals "PARITY-<runId> booking".',
    },
    '4c-item-reference': {
        'description': 'Web->mobile (API-set): inspection.itemReference PATCHed via the web API must '
                       'read back on GET /inspections and render on the mobile LocationCard (shown as '
                       'Asset Reference). Mirrors 2g/itemDetail.',
        'steps': [
            'API: PATCH /inspections/{id} itemReference = "PARITY-<runId> item-ref".',
            'Connection: GET /inspections/{id} -> .itemReference equals the set value (checkScalarField).',
            'Webapp: open the Inspections tab -> inspection; confirm the Asset Reference field.',
            'Mobile: open the inspection; read the LocationCard Asset Reference (p07, photo-only).',
        ],
        'expected': 'GET /inspections.itemReference equals the API-set value; mobile LocationCard shows it.',
        'connection': 'GET /inspections/{id} -> .itemReference equals "PARITY-<runId> item-ref".',
    },
    '4d-item-location': {
        'description': 'Web->mobile (API-set): inspection.itemLocation PATCHed via the web API must '
                       'read back on GET /inspections and render on the mobile LocationCard (shown as '
                       'Asset Location). Mirrors 2g/itemDetail.',
        'steps': [
            'API: PATCH /inspections/{id} itemLocation = "PARITY-<runId> item-loc".',
            'Connection: GET /inspections/{id} -> .itemLocation equals the set value (checkScalarField).',
            'Webapp: open the Inspections tab -> inspection; confirm the Asset Location field.',
            'Mobile: open the inspection; read the LocationCard Asset Location (p08, photo-only).',
        ],
        'expected': 'GET /inspections.itemLocation equals the API-set value; mobile LocationCard shows it.',
        'connection': 'GET /inspections/{id} -> .itemLocation equals "PARITY-<runId> item-loc".',
    },
    '2i-add-inspection': {
        'description': 'Web->mobile: a SECOND inspection added to the visit on the webapp (POST '
                       '/inspections with a DIFFERENT jobType than the primary Risk Assessment one) '
                       'must show on the visit -> inspections.length >= 2. Additive: never deletes or '
                       'modifies the first inspection (the existing checks depend on it).',
        'steps': [
            'API: GET /job-types; pick a jobType id != the primary (658f27c1...).',
            'API: POST /inspections {visitId, jobTypeId: <second jobType>} (the 1st inspection stays untouched).',
            'Connection: GET /visits/{id} -> .inspections.length >= 2 (checkInspectionCount).',
            'Webapp: open the visit -> Inspections tab; screenshot the list (now 2 inspections).',
            'Mobile: open the visit, tap the Inspections tab; screenshot the list (p10, photo-only).',
        ],
        'expected': 'GET /visits.inspections.length >= 2; the webapp Inspections tab lists 2 inspections.',
        'connection': 'GET /visits/{id} -> .inspections.length >= 2.',
    },
    '2j-visit-status': {
        'description': 'Web->mobile: the visit BOOKING status set on the webapp (PATCH /visits/{id} '
                       "status='confirmed', from the scheduled|pending|confirmed|cancelled enum) must "
                       'read back on GET /visits.status. GUARDRAIL: BOOKING status only -- never '
                       'touches visitStatus/inspectionStatus (the execution states). Probe-verified the '
                       "visit stays searchable with status='confirmed'.",
        'steps': [
            "API: PATCH /visits/{id} status = 'confirmed' (NOT 'cancelled', which can hide the visit).",
            "Connection: GET /visits/{id} -> .status equals 'confirmed' (checkScalarField).",
            'Webapp: open the visit details; screenshot the header status badge.',
            'Mobile: open the visit; screenshot the header status badge (p11, photo-only).',
        ],
        'expected': "GET /visits.status equals 'confirmed'; webapp header shows the Confirmed badge.",
        'connection': "GET /visits/{id} -> .status equals 'confirmed'.",
    },
    '4e-mobile-action': {
        'description': 'Mobile->web: a CUSTOM visit-level action ADDED ON MOBILE (the "add on mobile, '
                       'see on web" case) must read back via the web API. The mobile flow opens the '
                       'visit, opens the Actions FAB -> AddActionsBottomSheet, adds a custom action '
                       'named "PARITY-<runId> MobAct" with a High priority and Saves. setup-data NEVER '
                       'creates this action, so a PASS proves the mobile->web write actually happened. '
                       'DISTINCT "MobAct" name so it never collides with the 2b actions (Hi/Med/Lo).',
        'steps': [
            'Mobile (p12): open the visit; open the Quick-actions FAB -> "Actions"; tap "Add Custom Action".',
            'Mobile: type the action name "PARITY-<runId> MobAct"; set priority High; Save the action; Save the bottom sheet.',
            'Connection: GET /actions?visitId={id} -> an action named "PARITY-<runId> MobAct" is present (checkActionPresent).',
            'Webapp: open the visit Actions card; screenshot the new action row.',
        ],
        'expected': 'GET /actions?visitId contains an action named "PARITY-<runId> MobAct"; the webapp Actions card shows it.',
        'connection': 'GET /actions?visitId={id} -> some action .name == "PARITY-<runId> MobAct".',
    },
    '2k-sample-note': {
        'description': 'Web->mobile (API-set): a per-sample NOTE POSTed to the samples flagship\'s FIRST '
                       'laboratorySample (the inspection gets 16 samples via 2h/addSamples first) must '
                       'read back on GET /laboratory-samples/{id}.sampleNote.noteText and render on the '
                       'mobile Water Sampling sample. The note is NESTED on the sample (the flat .notes '
                       'field stays null). NEVER submits to Normec/ALS.',
        'steps': [
            'API (after 2h addSamples): GET /inspections/{id} -> laboratorySamples[0].id (persist the exact id).',
            'API: POST /laboratory-samples/{sampleId}/notes {noteText: "PARITY-<runId> sample-note"}.',
            'Connection: GET /laboratory-samples/{sampleId} -> .sampleNote.noteText equals the set value (checkSampleNote).',
            'Webapp: open the inspection -> Lab Results tab; expand the sample; confirm the note value.',
            'Mobile: open the inspection -> Water Sampling section; read the sample note (p13, photo-only).',
            'GUARDRAIL: never tap "Submit Samples" / POST /laboratory-samples/submit-batch.',
        ],
        'expected': 'GET /laboratory-samples.sampleNote.noteText equals the API-set value; the webapp Lab Results sample shows it.',
        'connection': 'GET /laboratory-samples/{sampleId} -> .sampleNote.noteText equals "PARITY-<runId> sample-note".',
    },
    '2l-engineers': {
        'description': 'Web->mobile: a SECOND engineer added to the visit on the webapp (PATCH '
                       '/visits/{id} {engineerIds:[existing, second]}) must show on the visit -> '
                       'visitEngineers.length >= 2. Write field is engineerIds; READ field is '
                       'visitEngineers (the drift note). GUARDRAIL: the existing engineer (parity.bot — '
                       'the mobile QA login) is KEPT structurally, so the visit never vanishes from mobile.',
        'steps': [
            'API: GET /visits/{id} -> current visitEngineers[].engineerId (the kept list).',
            'API: GET /users -> first isEngineer user not already on the visit (the 2nd engineer).',
            'API: PATCH /visits/{id} {engineerIds: [...existing, second]} (existing kept FIRST).',
            'Connection: GET /visits/{id} -> .visitEngineers.length >= 2 (checkEngineerCount).',
            'Webapp: open the visit details; screenshot the Engineers chips (now 2).',
            'Mobile: open the visit; screenshot the Engineers chips (p14, photo-only).',
        ],
        'expected': 'GET /visits.visitEngineers.length >= 2; the webapp Engineers chips show both engineers (parity.bot kept).',
        'connection': 'GET /visits/{id} -> .visitEngineers.length >= 2.',
    },
    '4f-ra-dropdowns': {
        'description': 'Web->mobile (API-set): the 36 Risk Assessment Yes/No DROPDOWN fields set via '
                       'PATCH /inspections/{id}/submit-form (the web UI renders them read-only per F-02) '
                       'must read back on GET /inspections and render in the mobile Risk Assessment '
                       "form's Yes/No dropdowns. ONE check, 36 datums (mirrors 2h = 16 samples). DISTINCT "
                       'from 3c (the 18 free-text "- Comments" of the same form). submit-form is MERGE, so '
                       'setting these never nulls the 3c comments. The single biggest dropdown-coverage '
                       'win: parity dropdown coverage 1 -> 37.',
        'steps': [
            'API: GET /inspections/{id} -> the Risk Assessment form\'s 36 dropdown fields (fieldOptions = Yes/No); resolve each InspectionFormField id by fieldName.',
            'API: PATCH /inspections/{id}/submit-form {formFields:[{id,value}]} with alternating Yes/No (one value per field).',
            'Connection: GET /inspections/{id} -> all 36 RA dropdown values equal the set values (checkFields ANDs every fieldName).',
            'Webapp: open the inspection -> Risk Assessment card; the 36 Yes/No selects are read-only and show the values.',
            'Mobile: open the inspection -> Risk Assessment form; the ExposedDropdownMenuBox dropdowns show the values (p15, photo-only — the render discriminator).',
        ],
        'expected': 'GET /inspections RA form: all 36 dropdowns equal the API-set Yes/No; mobile RA form renders them.',
        'connection': 'GET /inspections/{id} -> inspectionForms[formName="Risk Assessment"].formFields[36 dropdowns].value all match.',
    },
}


# ---------------------------------------------------------------------------
# openpyxl helpers (reused from the sibling generators)
# ---------------------------------------------------------------------------

def safe(value) -> str:
    if value is None:
        return ''
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return str(value)


def style_header(cell, fill='navy'):
    cell.fill = PatternFill('solid', fgColor=PALETTE[fill])
    cell.font = Font(name='Aptos', bold=True, size=11, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def card(ws, cell_range, title, value, fill_color, value_color='FFFFFF'):
    """Metric card — identical to generate_detector_excel.card()."""
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(':')[0]]
    cell.value = f'{title}\n{value}'
    cell.fill = PatternFill('solid', fgColor=fill_color)
    cell.font = Font(name='Aptos', bold=True, size=17, color=value_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color=PALETTE['border'])
    for row in ws[cell_range]:
        for item in row:
            item.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def add_image_scaled(ws, path_str, anchor_cell, max_w=240, max_h=140):
    """Embed an image scaled to fit max_w x max_h (same as webapp / detector)."""
    p = Path(path_str)
    if not p.exists():
        return False
    try:
        img = XLImage(str(p))
    except Exception:
        return False
    w, h = img.width, img.height
    ratio = min(max_w / max(w, 1), max_h / max(h, 1), 1)
    img.width = int(w * ratio)
    img.height = int(h * ratio)
    img.anchor = anchor_cell
    ws.add_image(img)
    return True


def status_colors(status_raw: str) -> tuple[str, str]:
    """(bg_fill, font_color) for a status cell."""
    s = (status_raw or '').upper()
    if s == 'PASS':
        return PALETTE['pass_bg'], PALETTE['pass']
    if s == 'SKIP':
        return PALETTE['skip_bg'], PALETTE['skip']
    return PALETTE['fail_bg'], PALETTE['fail']


def apply_status_fill(cell, status_raw: str):
    bg, fg = status_colors(status_raw)
    cell.fill = PatternFill('solid', fgColor=bg)
    cell.font = Font(name='Aptos', bold=True, color=fg)
    cell.alignment = Alignment(horizontal='center', vertical='center')


# ---------------------------------------------------------------------------
# Screenshot resolution
# ---------------------------------------------------------------------------

def find_mobile_after(screenshots_dir: Path, check_id: str) -> Path | None:
    """Return {flow}-after.png for the check, or None (incl. API-only 2c)."""
    flow = CHECK_TO_FLOW.get(check_id)
    if not flow:
        return None
    direct = screenshots_dir / f'{flow}-after.png'
    if direct.is_file():
        return direct
    # Defensive glob on the full prefix + trailing separator to avoid p03/p03b
    # collision (use the underscore that already separates the prefix from the
    # rest of the flow name — the full flow IS the prefix here, so anchor on it).
    matches = sorted(screenshots_dir.glob(f'{flow}-after.png'))
    return matches[0] if matches else None


def find_web_shot(screenshots_dir: Path, check_id: str, kind: str) -> Path | None:
    """Find an F2 web screenshot {check}-web-{kind}.png. kind in {set, verify}.

    Short check ids (2a, 2b, ...) do not collide, so a glob on the short token
    is safe. Tries short id, full id, then a short-prefix glob; first hit wins.
    """
    short = _short(check_id)
    candidates = [
        screenshots_dir / f'{short}-web-{kind}.png',
        screenshots_dir / f'{check_id}-web-{kind}.png',
    ]
    for c in candidates:
        if c.is_file():
            return c
    matches = sorted(screenshots_dir.glob(f'{short}*-web-{kind}.png'))
    return matches[0] if matches else None


# ---------------------------------------------------------------------------
# Actual-value rendering
# ---------------------------------------------------------------------------

def build_actual(check: dict) -> str:
    """Human "actual result" — prefer the structured fields dict over raw JSON."""
    fields = check.get('fields')
    details = safe(check.get('details', ''))
    if isinstance(fields, dict) and fields:
        ok = sum(1 for v in fields.values() if v)
        total = len(fields)
        names = ', '.join(k for k, v in fields.items() if v)
        head = f'{ok}/{total} fields verified'
        if names:
            head += f': {names}'
        return head
    # No structured fields — fall back to the details string (already human-ish
    # for 2a/2b/2c/2d/2g/3a).
    return details


# ---------------------------------------------------------------------------
# Sheet 1: Summary
# ---------------------------------------------------------------------------

def build_summary_sheet(ws, summary: dict, checks: list[dict],
                        screenshots_dir: Path, title: str):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90
    ws.sheet_properties.tabColor = '2563EB'

    run_id = safe(summary.get('runId', ''))
    visit_ref = safe(summary.get('visitRef', ''))
    total = int(summary.get('total', len(checks)) or 0)
    passed = int(summary.get('passed', 0) or 0)
    failed = int(summary.get('failed', 0) or 0)
    gate_failed = bool(summary.get('gateFailed', False))

    # Title banner
    ws.merge_cells('A1:G1')
    ws['A1'].value = title
    ws['A1'].fill = PatternFill('solid', fgColor=PALETTE['navy'])
    ws['A1'].font = Font(name='Aptos Display', bold=True, size=22, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 42

    # Subtitle
    ws.merge_cells('A2:G2')
    ws['A2'].value = f'Run {run_id}  |  Visit {visit_ref}  |  bidirectional web<->mobile parity'
    ws['A2'].fill = PatternFill('solid', fgColor=PALETTE['slate'])
    ws['A2'].font = Font(name='Aptos', size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 8

    # Metric cards — Total / Passed / Failed / Gate
    ws.row_dimensions[4].height = 52
    ws.row_dimensions[5].height = 52
    gate_label = 'FAILED' if gate_failed else 'PASS'
    gate_color = PALETTE['fail'] if gate_failed else PALETTE['pass']
    card(ws, 'A4:B5', 'Total', total, PALETTE['slate'])
    card(ws, 'C4:C5', 'Passed', passed, PALETTE['pass'])
    card(ws, 'D4:E5', 'Failed', failed, PALETTE['fail'])
    card(ws, 'F4:G5', 'Gate', gate_label, gate_color)

    ws.row_dimensions[6].height = 8
    ws.row_dimensions[7].height = 8

    # Table headers (row 8)
    headers = ['Check', 'Direction', 'Status', 'Connection (API)',
               'Web evidence', 'Mobile evidence', 'Detail']
    for i, h in enumerate(headers, start=1):
        style_header(ws.cell(row=8, column=i, value=h))
    ws.row_dimensions[8].height = 24

    row = 9
    for idx, check in enumerate(checks):
        check_id = safe(check.get('id', ''))
        direction = safe(check.get('direction', ''))
        status_raw = (check.get('status') or '').upper()
        meta = EVIDENCE_MAP.get(check_id, {})

        # Connection (API) — short form.
        connection = safe(meta.get('connection', ''))

        # Web evidence pointer (text only on Summary).
        if check_id == '2c-inspection-actions':
            web_ev = 'n/a (API-only, F-01)'
        elif find_web_shot(screenshots_dir, check_id, 'set') or \
                find_web_shot(screenshots_dir, check_id, 'verify'):
            web_ev = 'captured'
        else:
            web_ev = 'pending (F2)'

        # Mobile evidence pointer.
        if check_id == '2c-inspection-actions':
            mobile_ev = 'n/a (API-only, F-01)'
        elif find_mobile_after(screenshots_dir, check_id):
            mobile_ev = 'captured'
        else:
            mobile_ev = 'missing'

        ws.cell(row=row, column=1, value=check_id)
        ws.cell(row=row, column=2, value=direction)
        apply_status_fill(ws.cell(row=row, column=3, value=status_raw), status_raw)
        conn_cell = ws.cell(row=row, column=4, value=connection)
        conn_cell.alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=5, value=web_ev).alignment = Alignment(vertical='top')
        ws.cell(row=row, column=6, value=mobile_ev).alignment = Alignment(vertical='top')

        link_cell = ws.cell(row=row, column=7, value='See Evidence')
        # Evidence sheet keeps the same order, headers on row 1 -> data from row 2.
        link_cell.hyperlink = f"#'Evidence'!A{2 + idx}"
        link_cell.font = Font(name='Aptos', color='1D4ED8', underline='single')
        link_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[row].height = 40
        row += 1

    set_col_widths(ws, [22, 18, 10, 46, 16, 16, 14])
    ws.freeze_panes = 'A9'


# ---------------------------------------------------------------------------
# Sheet 2: Evidence (the proof sheet)
# ---------------------------------------------------------------------------

def build_evidence_sheet(ws, checks: list[dict], screenshots_dir: Path):
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 80
    ws.sheet_properties.tabColor = PALETTE['pass']

    headers = ['Check', 'Description', 'Steps to reproduce', 'Expected', 'Actual',
               'Connection check (API GET result)', 'Status',
               'Web screenshot', 'Mobile screenshot']
    for i, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=i, value=h))
    ws.row_dimensions[1].height = 24

    row = 2
    for check in checks:
        check_id = safe(check.get('id', ''))
        status_raw = (check.get('status') or '').upper()
        meta = EVIDENCE_MAP.get(check_id, {})

        description = safe(meta.get('description', ''))
        steps_list = meta.get('steps', []) or []
        steps = '\n'.join(f'{i}. {s}' for i, s in enumerate(steps_list, start=1))
        expected = safe(meta.get('expected', ''))
        actual = build_actual(check)
        connection = safe(meta.get('connection', ''))

        ws.cell(row=row, column=1, value=check_id).alignment = Alignment(vertical='top')
        ws.cell(row=row, column=2, value=description).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=3, value=steps).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=4, value=expected).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=5, value=actual).alignment = Alignment(wrap_text=True, vertical='top')
        ws.cell(row=row, column=6, value=connection).alignment = Alignment(wrap_text=True, vertical='top')
        apply_status_fill(ws.cell(row=row, column=7, value=status_raw), status_raw)

        ws.row_dimensions[row].height = 215  # ~ room for a 260px-scaled image

        # Web screenshot (column 8 / H) — F2 adds these; absent now -> placeholder.
        if check_id == '2c-inspection-actions':
            c = ws.cell(row=row, column=8, value='n/a - API-only (F-01)')
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.font = Font(name='Aptos', italic=True, size=9, color=PALETTE['muted'])
        else:
            web_path = (find_web_shot(screenshots_dir, check_id, 'set')
                        or find_web_shot(screenshots_dir, check_id, 'verify'))
            if web_path and add_image_scaled(ws, str(web_path), f'H{row}',
                                             max_w=260, max_h=260):
                pass
            else:
                c = ws.cell(row=row, column=8, value='web screenshot pending (F2)')
                c.alignment = Alignment(wrap_text=True, vertical='top')
                c.font = Font(name='Aptos', italic=True, size=9, color=PALETTE['muted'])

        # Mobile screenshot (column 9 / I) — {flow}-after.png.
        if check_id == '2c-inspection-actions':
            c = ws.cell(row=row, column=9, value='n/a - API-only mobile render gap (F-01)')
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.font = Font(name='Aptos', italic=True, size=9, color=PALETTE['muted'])
        else:
            mob_path = find_mobile_after(screenshots_dir, check_id)
            if mob_path and add_image_scaled(ws, str(mob_path), f'I{row}',
                                             max_w=260, max_h=260):
                pass
            else:
                c = ws.cell(row=row, column=9, value='mobile screenshot missing')
                c.alignment = Alignment(wrap_text=True, vertical='top')
                c.font = Font(name='Aptos', italic=True, size=9, color=PALETTE['muted'])

        row += 1

    set_col_widths(ws, [22, 44, 52, 40, 34, 46, 10, 40, 40])
    ws.freeze_panes = 'A2'


# ---------------------------------------------------------------------------
# Build + entrypoint
# ---------------------------------------------------------------------------

def load_summary(path: Path) -> dict:
    return json.loads(path.read_text(encoding='utf-8'))


def build_workbook(summary: dict, screenshots_dir: Path, output: Path,
                   title: str = 'Hydrocert Parity Evidence'):
    checks = list(summary.get('checks') or [])  # preserve JSON order (not numeric)

    wb = Workbook()
    wb.properties.title = title
    wb.properties.subject = 'Hydrocert bidirectional parity evidence'

    ws_summary = wb.active
    ws_summary.title = 'Summary'
    build_summary_sheet(ws_summary, summary, checks, screenshots_dir, title)

    ws_evidence = wb.create_sheet('Evidence')
    build_evidence_sheet(ws_evidence, checks, screenshots_dir)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output))
    return ws_evidence


def main(argv: list[str]) -> int:
    if len(argv) != 4:
        print('Usage: python build_parity_evidence_excel.py '
              '<summary.json> <screenshots_dir> <out.xlsx>', file=sys.stderr)
        return 2

    summary_path = Path(argv[1])
    screenshots_dir = Path(argv[2])
    output = Path(argv[3])

    if not summary_path.is_file():
        print(f'ERROR: summary JSON not found: {summary_path}', file=sys.stderr)
        return 1
    if not screenshots_dir.is_dir():
        print(f'WARNING: screenshots dir not found: {screenshots_dir} '
              '(images will degrade to placeholders)', file=sys.stderr)

    summary = load_summary(summary_path)
    build_workbook(summary, screenshots_dir, output)

    print(f'EXCEL_PATH={output}')
    return 0


if __name__ == '__main__':
    sys.exit(main(sys.argv))
