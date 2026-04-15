from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PALETTE = {
    'navy': '0F172A',
    'slate': '334155',
    'muted': '64748B',
    'border': 'CBD5E1',
    'pass': '0F766E',
    'pass_bg': 'CCFBF1',
    'fail': 'B91C1C',
    'fail_bg': 'FEE2E2',
    'skip': '92400E',
    'skip_bg': 'FEF3C7',
}

MAX_SHEET_NAME = 31
SHEET_NAME_RE = re.compile(r'[\\/*?:\[\]]')


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description='Generate UI Change Detector web-app Excel report.')
    p.add_argument('--diff-json', required=True)
    p.add_argument('--screenshots', required=True, help='Directory containing annotated/ and crops/ subdirs')
    p.add_argument('--output', required=True)
    return p.parse_args()


def safe_sheet_name(name: str, existing: set) -> str:
    base = SHEET_NAME_RE.sub('_', name).strip()[:MAX_SHEET_NAME] or 'Page'
    candidate = base
    idx = 2
    while candidate in existing:
        suffix = f' ({idx})'
        candidate = (base[: MAX_SHEET_NAME - len(suffix)] + suffix)
        idx += 1
    existing.add(candidate)
    return candidate


def style_header(cell, fill='navy'):
    cell.fill = PatternFill('solid', fgColor=PALETTE[fill])
    cell.font = Font(name='Aptos', bold=True, size=11, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_summary(ws, diff):
    totals = diff.get('totals', {})
    ws['A1'] = 'UI Change Detector — web-app'
    ws['A1'].font = Font(name='Aptos', bold=True, size=18, color=PALETTE['navy'])
    ws['A2'] = f"Mode: {diff.get('mode', '?')}  |  Routes: {totals.get('routes', 0)}  |  Missing: {totals.get('missing', 0)}  |  Introduced: {totals.get('introduced', 0)}  |  TextChanged: {totals.get('textChanged', 0)}  |  NewPages: {totals.get('newPages', 0)}  |  LostPages: {totals.get('lostPages', 0)}"
    ws['A2'].font = Font(name='Aptos', italic=True, size=10, color=PALETTE['slate'])

    headers = ['Page', 'URL', 'Missing', 'Introduced', 'TextChanged', 'Status']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    row = 5
    for path_key, data in diff.get('perPage', {}).items():
        missing = len(data.get('missing', []))
        introduced = len(data.get('introduced', []))
        text_changed = len(data.get('textChanged', []))
        if missing > 0:
            status = 'MISSING'
            status_fill = PALETTE['fail_bg']
            status_color = PALETTE['fail']
        elif introduced > 0 or text_changed > 0:
            status = 'INTRODUCED'
            status_fill = PALETTE['skip_bg']
            status_color = PALETTE['skip']
        else:
            status = 'OK'
            status_fill = PALETTE['pass_bg']
            status_color = PALETTE['pass']

        ws.cell(row=row, column=1, value=path_key)
        ws.cell(row=row, column=2, value=data.get('url', ''))
        ws.cell(row=row, column=3, value=missing)
        ws.cell(row=row, column=4, value=introduced)
        ws.cell(row=row, column=5, value=text_changed)
        status_cell = ws.cell(row=row, column=6, value=status)
        status_cell.fill = PatternFill('solid', fgColor=status_fill)
        status_cell.font = Font(name='Aptos', bold=True, color=status_color)
        status_cell.alignment = Alignment(horizontal='center')
        row += 1

    for p in diff.get('newPages', []):
        ws.cell(row=row, column=1, value=p)
        ws.cell(row=row, column=6, value='NEW PAGE').fill = PatternFill('solid', fgColor=PALETTE['skip_bg'])
        row += 1
    for p in diff.get('lostPages', []):
        ws.cell(row=row, column=1, value=p)
        ws.cell(row=row, column=6, value='LOST PAGE').fill = PatternFill('solid', fgColor=PALETTE['fail_bg'])
        row += 1

    set_col_widths(ws, [36, 48, 10, 12, 14, 14])


def add_image_scaled(ws, path_str, anchor_cell, max_width_px=240, max_height_px=140):
    p = Path(path_str)
    if not p.exists():
        return
    try:
        img = XLImage(str(p))
    except Exception:
        return
    w, h = img.width, img.height
    ratio_w = max_width_px / max(w, 1)
    ratio_h = max_height_px / max(h, 1)
    ratio = min(ratio_w, ratio_h, 1)
    img.width = int(w * ratio)
    img.height = int(h * ratio)
    img.anchor = anchor_cell
    ws.add_image(img)


def write_page_sheet(ws, path_key, data):
    ws['A1'] = f'Route: {path_key}'
    ws['A1'].font = Font(name='Aptos', bold=True, size=14, color=PALETTE['navy'])
    ws['A2'] = data.get('url', '')
    ws['A2'].font = Font(name='Aptos', italic=True, size=10, color=PALETTE['muted'])

    headers = ['Change Type', 'Role', 'Accessible Name', 'Text', 'Selector Hint', 'BBox (x,y,w,h)', 'Print-screen']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    row = 5
    missing_crops = data.get('missingCrops', [])
    introduced_crops = data.get('introducedCrops', [])

    for idx, el in enumerate(data.get('missing', [])):
        ws.cell(row=row, column=1, value='Missing').fill = PatternFill('solid', fgColor=PALETTE['fail_bg'])
        ws.cell(row=row, column=2, value=el.get('role', ''))
        ws.cell(row=row, column=3, value=el.get('name', ''))
        ws.cell(row=row, column=4, value=(el.get('text') or '')[:200])
        ws.cell(row=row, column=5, value=el.get('selectorHint', ''))
        b = el.get('bbox') or {}
        ws.cell(row=row, column=6, value=f"{b.get('x','')},{b.get('y','')},{b.get('w','')},{b.get('h','')}")
        ws.row_dimensions[row].height = 110
        crop = missing_crops[idx] if idx < len(missing_crops) else None
        if crop:
            add_image_scaled(ws, crop, f'G{row}')
        row += 1

    for idx, el in enumerate(data.get('introduced', [])):
        ws.cell(row=row, column=1, value='Introduced').fill = PatternFill('solid', fgColor=PALETTE['skip_bg'])
        ws.cell(row=row, column=2, value=el.get('role', ''))
        ws.cell(row=row, column=3, value=el.get('name', ''))
        ws.cell(row=row, column=4, value=(el.get('text') or '')[:200])
        ws.cell(row=row, column=5, value=el.get('selectorHint', ''))
        b = el.get('bbox') or {}
        ws.cell(row=row, column=6, value=f"{b.get('x','')},{b.get('y','')},{b.get('w','')},{b.get('h','')}")
        ws.row_dimensions[row].height = 110
        crop = introduced_crops[idx] if idx < len(introduced_crops) else None
        if crop:
            add_image_scaled(ws, crop, f'G{row}')
        row += 1

    for el in data.get('textChanged', []):
        ws.cell(row=row, column=1, value='TextChanged').fill = PatternFill('solid', fgColor=PALETTE['skip_bg'])
        ws.cell(row=row, column=2, value=el.get('role', ''))
        ws.cell(row=row, column=3, value=el.get('name', ''))
        ws.cell(row=row, column=4, value=f"{(el.get('baselineText') or '')[:90]} → {(el.get('currentText') or '')[:90]}")
        ws.cell(row=row, column=5, value=el.get('selectorHint', ''))
        b = el.get('bbox') or {}
        ws.cell(row=row, column=6, value=f"{b.get('x','')},{b.get('y','')},{b.get('w','')},{b.get('h','')}")
        row += 1

    row += 2
    cur_ann = data.get('currentAnnotated')
    base_ann = data.get('baselineAnnotated')
    if base_ann:
        ws.cell(row=row, column=1, value='Baseline (with Missing circled):').font = Font(bold=True)
        add_image_scaled(ws, base_ann, f'A{row + 1}', max_width_px=520, max_height_px=1000)
    if cur_ann:
        ws.cell(row=row, column=4, value='Current (with Introduced circled):').font = Font(bold=True)
        add_image_scaled(ws, cur_ann, f'D{row + 1}', max_width_px=520, max_height_px=1000)

    set_col_widths(ws, [14, 14, 32, 36, 36, 20, 40])


def main():
    args = parse_args()
    diff = json.loads(Path(args.diff_json).read_text(encoding='utf-8'))

    wb = Workbook()
    wb.remove(wb.active)
    ws_summary = wb.create_sheet(title='Summary')
    write_summary(ws_summary, diff)

    existing_names = {'Summary'}
    for path_key, data in diff.get('perPage', {}).items():
        missing = data.get('missing', [])
        introduced = data.get('introduced', [])
        text_changed = data.get('textChanged', [])
        if not (missing or introduced or text_changed):
            continue
        name = safe_sheet_name(path_key.replace('/', ' ').strip() or 'root', existing_names)
        ws = wb.create_sheet(title=name)
        write_page_sheet(ws, path_key, data)

    output = Path(args.output)
    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    print(f'Wrote {output}')


if __name__ == '__main__':
    main()
