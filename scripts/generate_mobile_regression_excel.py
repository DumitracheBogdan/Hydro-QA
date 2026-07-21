"""
Generate an Excel report for the mobile Maestro post-deploy regression workflow.

Consumes:
  - summary.json produced by scripts/run-mobile-v2-test.sh
  - Per-flow after.png screenshots under <artifacts>/test/screenshots/
  - Per-flow uiautomator XML dumps under <artifacts>/test/ui-dumps/
  - Per-flow maestro logs under <artifacts>/logs/ (or test/logs/)
  - Per-flow Maestro YAML files under --flows-dir (default: mobile-flows-v2) —
    the leading `# ...` comment block after the `---` front-matter separator
    is extracted as a human-readable "What it checks" description column.

Produces a workbook with TWO sheets modeled on generate_detector_excel.py:

  Sheet 1 "Summary" — title banner, subtitle, 4 metric cards (Total / Passed /
  Failed / Skipped), and a per-flow table (Flow | What it checks | Status |
  Duration | Error | Detail). The Detail column is a cross-sheet hyperlink to
  the matching row on the Details sheet.

  Sheet 2 "Details" — per-flow rich view with the current annotated-screenshot
  layout (Flow | What it checks | Status | Error | Print-screen (annotated)).

For failed flows, attempts to parse the selector text out of the Maestro error
line, find a matching node in the uiautomator dump, and draw a red circle +
numbered label on the after screenshot (reusing the pattern from
generate_detector_excel.py). If nothing matches, the raw after.png is embedded
instead — the script never crashes on missing pieces.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import tempfile
import xml.etree.ElementTree as ET
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# Reuse the annotate + styling helpers from the sibling detector script if importable.
sys.path.insert(0, str(Path(__file__).resolve().parent))
try:
    from generate_detector_excel import (  # type: ignore
        PALETTE,
        parse_bounds,
        crop_and_annotate,
        add_image_scaled,
        card as _detector_card,
        style_header as _detector_style_header,
    )
    _HAS_DETECTOR_HELPERS = True
except Exception:
    _HAS_DETECTOR_HELPERS = False
    # Fallback: copy minimal pieces so we don't hard-depend on the sibling.
    PALETTE = {
        'navy': '0F172A',
        'slate': '334155',
        'muted': '64748B',
        'border': 'CBD5E1',
        'pass': '0F766E',
        'pass_bg': 'CCFBF1',
        'fail': 'B91C1C',
        'fail_bg': 'FEE2E2',
        'skip': '92400E',
        'skip_bg': 'FEF3C7',
    }
    _BOUNDS_RE = re.compile(r'\[(\d+),(\d+)\]\[(\d+),(\d+)\]')

    def parse_bounds(s):  # type: ignore[override]
        m = _BOUNDS_RE.match(s or '')
        if not m:
            return None
        return int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))

    def crop_and_annotate(*args, **kwargs):  # type: ignore[override]
        return None

    def add_image_scaled(ws, path_str, anchor_cell, max_w=360, max_h=520):  # type: ignore[override]
        p = Path(path_str)
        if not p.exists():
            return
        try:
            img = XLImage(str(p))
        except Exception:
            return
        w, h = img.width, img.height
        ratio = min(max_w / max(w, 1), max_h / max(h, 1), 1)
        img.width = int(w * ratio)
        img.height = int(h * ratio)
        img.anchor = anchor_cell
        ws.add_image(img)


# ---------------------------------------------------------------------------
# Local styling helpers (fallback if the detector helpers weren't importable)
# ---------------------------------------------------------------------------

def style_header(cell, fill='navy'):
    """Navy header style — delegates to the detector helper when available."""
    if _HAS_DETECTOR_HELPERS:
        return _detector_style_header(cell, fill=fill)
    cell.fill = PatternFill('solid', fgColor=PALETTE[fill])
    cell.font = Font(name='Aptos', bold=True, size=11, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def card(ws, cell_range, title, value, fill_color, value_color='FFFFFF'):
    """Metric card — delegates to the detector helper when available."""
    if _HAS_DETECTOR_HELPERS:
        return _detector_card(ws, cell_range, title, value, fill_color, value_color=value_color)
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(':')[0]]
    cell.value = f'{title}\n{value}'
    cell.fill = PatternFill('solid', fgColor=fill_color)
    cell.font = Font(name='Aptos', bold=True, size=17, color=value_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color=PALETTE.get('border', 'CBD5E1'))
    for row in ws[cell_range]:
        for item in row:
            item.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


# Quoted-selector regex: e.g.  Element "Save" not found, or 'Save'
QUOTED_RE = re.compile(r'["\u201C\u201D\u2018\u2019\'`]([^"\u201C\u201D\u2018\u2019\'`]{1,120})["\u201C\u201D\u2018\u2019\'`]')


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description='Generate mobile regression Excel report.')
    p.add_argument('--summary-json', required=True, help='Path to summary.json from run-mobile-v2-test.sh')
    p.add_argument('--artifacts-dir', required=True, help='Path to qa-artifacts/mobile-v2')
    p.add_argument('--output', required=True, help='Output .xlsx path')
    p.add_argument('--title', default='Mobile Regression Report')
    p.add_argument('--subtitle', default='')
    p.add_argument(
        '--flows-dir',
        default='mobile-flows-v2',
        help='Directory containing the Maestro YAML flows (used to extract the '
             '"What it checks" description column). Relative to cwd or absolute.',
    )
    return p.parse_args()


def extract_flow_description(flows_dir: Path, flow_id: str) -> str:
    """Return the leading `#` comment block from a Maestro YAML flow.

    Opens {flows_dir}/{flow_id}.yaml and looks for the `---` front-matter
    separator. After it, skips blank lines, then reads consecutive `#` comment
    lines (stripping the leading `# ` / `#`). Lines starting with "Source:"
    (case-insensitive, after strip) are skipped. The first non-comment line
    terminates the block. Returns '' if the file is missing, unreadable, or
    has no usable comment block. Output is truncated at 280 chars with `…`.
    """
    path = flows_dir / f'{flow_id}.yaml'
    try:
        text = path.read_text(encoding='utf-8', errors='replace')
    except Exception:
        return ''

    lines = text.splitlines()

    # Find the --- front-matter separator.
    sep_idx = -1
    for i, line in enumerate(lines):
        if line.strip() == '---':
            sep_idx = i
            break
    if sep_idx < 0:
        return ''

    # Skip blank lines after the separator.
    i = sep_idx + 1
    while i < len(lines) and not lines[i].strip():
        i += 1

    # Read consecutive `#` comment lines.
    cleaned: list[str] = []
    while i < len(lines):
        raw = lines[i]
        stripped = raw.lstrip()
        if not stripped.startswith('#'):
            break
        # Strip leading '#' then an optional single space.
        body = stripped[1:]
        if body.startswith(' '):
            body = body[1:]
        body = body.strip()
        # Skip doc-reference lines: "Source:", "Source of truth:", etc.
        if body and not re.match(r'(?i)source\b', body):
            cleaned.append(body)
        i += 1

    if not cleaned:
        return ''

    joined = ' '.join(cleaned)
    joined = re.sub(r'\s+', ' ', joined).strip()
    if len(joined) > 280:
        joined = joined[:279].rstrip() + '\u2026'
    return joined


def load_summary(path: Path) -> dict:
    try:
        return json.loads(path.read_text(encoding='utf-8'))
    except Exception as exc:
        print(f'WARNING: failed to load summary {path}: {exc}', file=sys.stderr)
        return {'checks': [], 'totals': {}}


def find_log_line(log_path: Path) -> str:
    """Return the first line matching FAILED|Error|Assertion is false, or ''."""
    if not log_path.is_file():
        return ''
    try:
        text = log_path.read_text(encoding='utf-8', errors='replace')
    except Exception:
        return ''
    for line in text.splitlines():
        if re.search(r'FAILED|Assertion is false|Error', line, re.IGNORECASE):
            return line.strip()[:400]
    return ''


def extract_selector(error_line: str) -> str | None:
    """Best-effort: pull a quoted selector string out of the error line."""
    if not error_line:
        return None
    candidates = [c for c in QUOTED_RE.findall(error_line) if c.strip()]
    if not candidates:
        return None
    candidates.sort(key=len, reverse=True)
    return candidates[0].strip()


def find_node_bounds(xml_path: Path, selector: str) -> tuple[int, int, int, int] | None:
    """Search a uiautomator XML dump for a node whose text / content-desc matches."""
    if not xml_path.is_file() or not selector:
        return None
    try:
        tree = ET.parse(str(xml_path))
    except Exception as exc:
        print(f'WARNING: failed to parse {xml_path}: {exc}', file=sys.stderr)
        return None

    root = tree.getroot()
    needle = selector.lower()

    exact_attrs = ('text', 'content-desc', 'resource-id')
    # First pass: exact match
    for node in root.iter('node'):
        for attr in exact_attrs:
            v = node.attrib.get(attr, '')
            if v and v.lower() == needle:
                b = parse_bounds(node.attrib.get('bounds', ''))
                if b:
                    return b
    # Second pass: substring match
    for node in root.iter('node'):
        for attr in exact_attrs:
            v = node.attrib.get(attr, '')
            if v and needle in v.lower():
                b = parse_bounds(node.attrib.get('bounds', ''))
                if b:
                    return b
    return None


# Selectors that are navigation / login / cleanup scaffolding, never the
# element under test. Used when a flow has no explicit '# CIRCLE:' hint.
_GENERIC_SELECTORS = {
    'history', 'activity', 'visits', 'account', 'back', 'not now',
    'type to search', 'overview of your visits', 'welcome back', 'login',
    'email', 'password', 'apply', 'development', 'hydrocert logo',
    'select environment', 'view visit details', 'inspections', 'quick actions',
    'forgot your password?',
}
_SELECTOR_LINE_RE = re.compile(r'(?:tapOn|assertVisible|text)\s*:\s*["\']([^"\']+)["\']')
_CIRCLE_HINT_RE = re.compile(r'#\s*CIRCLE:\s*(.+?)\s*$', re.MULTILINE)


def _clean_selector(sel: str) -> str:
    """Turn a Maestro regex selector into a plain substring for matching."""
    s = (sel or '').strip().strip('"\'')
    if s.startswith('(') and '|' in s:            # (A|B) -> A
        s = s[1:].split('|', 1)[0]
    for ch in ('^', '$', '(', ')', '\\'):
        s = s.replace(ch, '')
    s = s.replace('.*', '').replace('d+', '').replace('dd', '')
    return re.sub(r'\s+', ' ', s).strip()


def derive_circle_targets(flows_dir: Path, flow_id: str) -> list:
    """The elements to circle for this flow, from an explicit
    '# CIRCLE: A, B, C' hint (comma-separated). Each becomes its own numbered
    circle so a test that verifies several things shows every one of them.
    The hint is the single source of truth - a flow with no hint gets no
    circle (raw screenshot) rather than a guessed, possibly-wrong one."""
    p = flows_dir / f'{flow_id}.yaml'
    if not p.is_file():
        return []
    text = p.read_text(encoding='utf-8', errors='ignore')
    m = _CIRCLE_HINT_RE.search(text)
    if not m:
        return []
    out = []
    for part in m.group(1).split(','):
        c = _clean_selector(part)
        if c and c not in out:
            out.append(c)
    return out


def annotate_full(img_path: Path, bounds_list, tmp_dir: Path, tag: str):
    """Draw a red ellipse + numbered badge (1, 2, 3 ...) around EACH element in
    `bounds_list` on the FULL screenshot, keeping on-screen context, so the dev
    sees every element the test verified. Returns the annotated copy path."""
    if not HAS_PIL or not bounds_list:
        return None
    try:
        img = PILImage.open(img_path).convert('RGB')
    except Exception:
        return None
    draw = ImageDraw.Draw(img)
    r = max(12, int(img.width * 0.024))
    line_w = max(4, int(img.width * 0.008))
    try:
        font = ImageFont.truetype('arialbd.ttf', int(r * 1.3))
    except Exception:
        font = ImageFont.load_default()
    for i, bounds in enumerate(bounds_list, 1):
        x1, y1, x2, y2 = bounds
        pad_x = max(14, int((x2 - x1) * 0.18))
        pad_y = max(14, int((y2 - y1) * 0.35))
        ex1, ey1 = max(0, x1 - pad_x), max(0, y1 - pad_y)
        ex2, ey2 = min(img.width, x2 + pad_x), min(img.height, y2 + pad_y)
        draw.ellipse([ex1, ey1, ex2, ey2], outline=(214, 40, 40), width=line_w)
        bx = min(img.width - 2 * r, ex2 - r)
        by = max(0, ey1 - r)
        draw.ellipse([bx, by, bx + 2 * r, by + 2 * r], fill=(214, 40, 40))
        s = str(i)
        tw = draw.textlength(s, font=font)
        draw.text((bx + r - tw / 2, by + r * 0.25), s, fill='white', font=font)
    out = tmp_dir / f'{tag}_full.png'
    try:
        img.save(out)
    except Exception:
        return None
    return out


def flow_to_steps(flows_dir: Path, flow_id: str, max_steps: int = 14) -> list:
    """Translate a flow file into simple, human 'steps to reproduce' - the kind
    a person can follow by hand: Log in / Open X / Tap "Y" / Type "Z" / See "W".
    Skips the technical noise (waits, timeouts, hideKeyboard, screenshots)."""
    p = flows_dir / f'{flow_id}.yaml'
    if not p.is_file():
        return []
    lines = p.read_text(encoding='utf-8', errors='ignore').split('\n')
    start = 0
    for i, ln in enumerate(lines):
        if ln.strip() == '---':
            start = i + 1
            break

    def nearby_text(j, window=5):
        for k in range(j, min(j + window, len(lines))):
            m = re.search(r'text:\s*["\']([^"\']+)', lines[k])
            if m:
                return _clean_selector(m.group(1))
        return None

    steps: list[str] = []
    i = start
    while i < len(lines):
        s = lines[i].strip()
        if not s or s.startswith('#'):
            i += 1
            continue
        m = re.match(r'-\s*runFlow:\s*(\S+)', s)
        if m:
            f = m.group(1)
            if 'login' in f:
                steps.append('Log in')
            elif 'open_qa_test' in f:
                steps.append('Open the "QA test" visit (History > search "QA test" > View Visit Details)')
            elif 'open_qa_forms' in f:
                steps.append('Open the "QA forms" visit from History')
            elif 'open_qa_procdeath' in f:
                steps.append('Open the "QA procdeath" visit from History')
            i += 1
            continue
        if re.match(r'-\s*launchApp', s):
            steps.append('Open the app')
            i += 1
            continue
        m = re.match(r'-\s*tapOn:\s*["\']([^"\']+)', s)
        if m:
            steps.append(f'Tap "{_clean_selector(m.group(1))}"')
            i += 1
            continue
        if re.match(r'-\s*tapOn:\s*$', s):
            t = nearby_text(i + 1)
            if t:
                steps.append(f'Tap "{t}"')
            i += 1
            continue
        m = re.match(r'-\s*inputText:\s*["\']?([^"\'\n]+)', s)
        if m:
            v = m.group(1).strip()
            if v.startswith('${'):
                v = 'your email' if 'EMAIL' in v else 'your password' if 'PASSWORD' in v else 'text'
                steps.append(f'Type {v}')
            else:
                steps.append(f'Type "{v}"')
            i += 1
            continue
        m = re.match(r'-\s*assertVisible:\s*["\']([^"\']+)', s)
        if m:
            steps.append(f'See "{_clean_selector(m.group(1))}"')
            i += 1
            continue
        m = re.match(r'-\s*assertNotVisible:\s*["\']([^"\']+)', s)
        if m:
            steps.append(f'Confirm "{_clean_selector(m.group(1))}" is gone')
            i += 1
            continue
        if re.match(r'-\s*scrollUntilVisible', s):
            t = nearby_text(i + 1, 6)
            if t:
                steps.append(f'Scroll to "{t}"')
            i += 1
            continue
        if s == '- back':
            steps.append('Press back')
            i += 1
            continue
        i += 1

    out: list[str] = []
    for st in steps:
        if not out or out[-1] != st:
            out.append(st)
    return out[:max_steps]


def resolve_log_path(artifacts_dir: Path, flow_name: str) -> Path:
    """Logs may be in artifacts/logs/ or artifacts/test/logs/ — try both."""
    candidates = [
        artifacts_dir / 'test' / 'logs' / f'{flow_name}.log',
        artifacts_dir / 'logs' / f'{flow_name}.log',
    ]
    for c in candidates:
        if c.is_file():
            return c
    return candidates[0]


def _status_colors(status_raw: str) -> tuple[str, str]:
    """Return (bg_fill, font_color) for a status cell."""
    if status_raw == 'PASS':
        return PALETTE['pass_bg'], PALETTE['pass']
    if status_raw == 'SKIP':
        return PALETTE.get('skip_bg', 'FEF3C7'), PALETTE.get('skip', '92400E')
    return PALETTE['fail_bg'], PALETTE['fail']


# ===================================================================
# Sheet 1: Summary
# ===================================================================

def build_summary_sheet(
    ws,
    summary: dict,
    title: str,
    subtitle: str,
    descriptions: dict[str, str] | None = None,
) -> None:
    """Title banner + 4 metric cards + per-flow table with cross-sheet links."""
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90

    descriptions = descriptions or {}

    totals = summary.get('totals') or {}
    generated = summary.get('generatedAt', '')
    total = int(totals.get('total', 0) or 0)
    passed = int(totals.get('pass', 0) or 0)
    failed = int(totals.get('fail', 0) or 0)
    skipped = int(totals.get('skip', 0) or 0)

    # Row 1: Title banner — keep 8-col span (cards still anchor A:H).
    ws.merge_cells('A1:H1')
    ws['A1'].value = title
    ws['A1'].fill = PatternFill('solid', fgColor=PALETTE['navy'])
    ws['A1'].font = Font(name='Aptos Display', bold=True, size=22, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 42

    # Row 2: Subtitle
    sub_text = subtitle if subtitle else f'Generated: {generated}'
    ws.merge_cells('A2:H2')
    ws['A2'].value = sub_text
    ws['A2'].fill = PatternFill('solid', fgColor=PALETTE['slate'])
    ws['A2'].font = Font(name='Aptos', size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 28

    # Row 3: spacer
    ws.row_dimensions[3].height = 8

    # Rows 4-5: Metric cards
    ws.row_dimensions[4].height = 52
    ws.row_dimensions[5].height = 52
    card(ws, 'A4:B5', 'Total',   total,   PALETTE['slate'])
    card(ws, 'C4:D5', 'Passed',  passed,  PALETTE['pass'])
    card(ws, 'E4:F5', 'Failed',  failed,  PALETTE['fail'])
    card(ws, 'G4:H5', 'Skipped', skipped, PALETTE.get('skip', '92400E'), value_color='000000')

    # Rows 6-7: spacers
    ws.row_dimensions[6].height = 8
    ws.row_dimensions[7].height = 8

    # Row 8: table headers
    # Columns: Flow=1, What it checks=2, Status=3, Duration=4, Error=5, Detail=6
    headers = ['Flow', 'What it checks', 'Status', 'Duration', 'Error', 'Detail']
    for i, h in enumerate(headers, start=1):
        style_header(ws.cell(row=8, column=i, value=h))
    ws.row_dimensions[8].height = 24

    # Row 9+: one row per check. Details sheet will have the same order
    # starting at row 2 (row 1 = headers), so details_row = 2 + index.
    checks = summary.get('checks') or []
    row = 9
    for idx, check in enumerate(checks):
        flow_name = check.get('id') or ''
        status_raw = (check.get('status') or '').upper()
        duration = check.get('duration', '') or check.get('durationMs', '') or ''
        details = check.get('details') or ''
        description = descriptions.get(flow_name, '') or ''

        ws.cell(row=row, column=1, value=flow_name)

        what_cell = ws.cell(row=row, column=2, value=description)
        what_cell.alignment = Alignment(wrap_text=True, vertical='top')

        sc = ws.cell(row=row, column=3, value=status_raw)
        bg, fg = _status_colors(status_raw)
        sc.fill = PatternFill('solid', fgColor=bg)
        sc.font = Font(name='Aptos', bold=True, color=fg)
        sc.alignment = Alignment(horizontal='center', vertical='center')

        ws.cell(row=row, column=4, value=duration if duration != '' else None)

        err_cell = ws.cell(row=row, column=5, value=details[:80])
        err_cell.alignment = Alignment(wrap_text=True, vertical='top')

        link_cell = ws.cell(row=row, column=6, value='See Details')
        details_target_row = 2 + idx  # headers on row 1
        link_cell.hyperlink = f"#'Details'!A{details_target_row}"
        link_cell.font = Font(name='Aptos', color='1D4ED8', underline='single')
        link_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[row].height = 60

        row += 1

    set_col_widths(ws, [28, 50, 10, 12, 55, 15])

    # Freeze header + left column
    ws.freeze_panes = 'B9'


# ===================================================================
# Sheet 2: Details
# ===================================================================

def build_details_sheet(
    ws,
    summary: dict,
    artifacts_dir: Path,
    tmp_dir: Path,
    descriptions: dict[str, str] | None = None,
    circle_targets: dict[str, str] | None = None,
    steps: dict[str, list] | None = None,
) -> None:
    """Per-flow rich view: Flow | What it checks | Steps to reproduce | Status |
    Error | Print-screen (annotated with the tested elements circled)."""
    ws.sheet_view.showGridLines = False

    descriptions = descriptions or {}
    circle_targets = circle_targets or {}
    steps = steps or {}

    # Row 1: headers (no title banner, Summary sheet already has it)
    headers = ['Flow', 'What it checks', 'Steps to reproduce', 'Status', 'Error',
               'Print-screen (annotated)']
    for i, h in enumerate(headers, start=1):
        style_header(ws.cell(row=1, column=i, value=h))
    ws.row_dimensions[1].height = 24

    screenshots_dir = artifacts_dir / 'test' / 'screenshots'
    ui_dumps_dir = artifacts_dir / 'test' / 'ui-dumps'

    checks = summary.get('checks') or []

    row = 2
    for check in checks:
        flow_name = check.get('id') or ''
        status_raw = (check.get('status') or '').upper()
        details = check.get('details') or ''
        description = descriptions.get(flow_name, '') or ''

        ws.cell(row=row, column=1, value=flow_name)

        what_cell = ws.cell(row=row, column=2, value=description)
        what_cell.alignment = Alignment(wrap_text=True, vertical='top')

        # Steps to reproduce - simple numbered human steps.
        flow_steps = steps.get(flow_name) or []
        steps_text = '\n'.join(f'{i}. {s}' for i, s in enumerate(flow_steps, 1)) or '-'
        steps_cell = ws.cell(row=row, column=3, value=steps_text)
        steps_cell.alignment = Alignment(wrap_text=True, vertical='top')

        status_cell = ws.cell(row=row, column=4, value=status_raw)
        bg, fg = _status_colors(status_raw)
        status_cell.fill = PatternFill('solid', fgColor=bg)
        status_cell.font = Font(name='Aptos', bold=True, color=fg)
        status_cell.alignment = Alignment(horizontal='center', vertical='center')

        # For failures, enrich error with the log line.
        error_text = details
        if status_raw != 'PASS':
            log_line = find_log_line(resolve_log_path(artifacts_dir, flow_name))
            if log_line and log_line not in error_text:
                error_text = (error_text + '  |  ' + log_line).strip(' |')

        err_cell = ws.cell(row=row, column=5, value=error_text[:800])
        err_cell.alignment = Alignment(wrap_text=True, vertical='top')

        screenshot_path = screenshots_dir / f'{flow_name}-after.png'
        ws.row_dimensions[row].height = 320
        embed_path: Path | None = None
        dump = ui_dumps_dir / f'{flow_name}.xml'

        # Annotation - the SAME "what was tested" circles for pass and fail:
        # the flow's '# CIRCLE:' elements, each drawn as a numbered circle on
        # the full screenshot. On a failure the same circles show what the test
        # was verifying; if none resolve, fall back to circling the element
        # named in the failure, else the raw screenshot.
        if HAS_PIL and screenshot_path.is_file():
            targets = circle_targets.get(flow_name) or []
            if isinstance(targets, str):
                targets = [targets]
            bounds_list = []
            for t in targets:
                b = find_node_bounds(dump, t)
                if b is not None and b not in bounds_list:
                    bounds_list.append(b)
            if bounds_list:
                try:
                    annotated = annotate_full(screenshot_path, bounds_list, tmp_dir,
                                              tag=f'{flow_name}_ann')
                    if annotated:
                        embed_path = annotated
                except Exception as exc:
                    print(f'WARNING: annotate failed for {flow_name}: {exc}', file=sys.stderr)
            if embed_path is None and status_raw != 'PASS':
                selector = extract_selector(error_text)
                b = find_node_bounds(dump, selector) if selector else None
                if b is not None:
                    try:
                        img = PILImage.open(screenshot_path)
                        annotated = crop_and_annotate(img, b[0], b[1], b[2], b[3],
                                                      label=1, tmp_dir=tmp_dir, tag=f'{flow_name}_fail')
                        if annotated:
                            embed_path = annotated
                    except Exception as exc:
                        print(f'WARNING: fail-annotate for {flow_name}: {exc}', file=sys.stderr)
            if embed_path is None:
                embed_path = screenshot_path

        if embed_path is not None:
            try:
                add_image_scaled(ws, str(embed_path), f'F{row}', max_w=360, max_h=520)
            except Exception as exc:
                print(f'WARNING: embed image failed for {flow_name}: {exc}', file=sys.stderr)

        row += 1

    set_col_widths(ws, [22, 40, 46, 10, 44, 52])
    ws.freeze_panes = 'A2'


# ===================================================================
# Build + entrypoint
# ===================================================================

def build_report(
    summary: dict,
    artifacts_dir: Path,
    output: Path,
    title: str,
    subtitle: str,
    descriptions: dict[str, str] | None = None,
    circle_targets: dict[str, str] | None = None,
    steps: dict[str, list] | None = None,
) -> None:
    wb = Workbook()
    descriptions = descriptions or {}
    circle_targets = circle_targets or {}
    steps = steps or {}

    with tempfile.TemporaryDirectory(prefix='mobile_excel_') as tmp_dir_str:
        tmp_dir = Path(tmp_dir_str)

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        build_summary_sheet(ws_summary, summary, title, subtitle, descriptions)

        # Sheet 2: Details
        ws_details = wb.create_sheet(title='Details')
        build_details_sheet(ws_details, summary, artifacts_dir, tmp_dir, descriptions, circle_targets, steps)

        output.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output))


def main() -> int:
    args = parse_args()
    summary_path = Path(args.summary_json)
    artifacts_dir = Path(args.artifacts_dir)
    output = Path(args.output)
    flows_dir = Path(args.flows_dir)

    if not HAS_PIL:
        print('WARNING: Pillow not installed — screenshots will be embedded unannotated', file=sys.stderr)

    if not summary_path.is_file():
        print(f'WARNING: summary JSON not found at {summary_path} — writing empty report', file=sys.stderr)
        summary = {'checks': [], 'totals': {}}
    else:
        summary = load_summary(summary_path)

    # Pre-compute per flow: the "What it checks" description and the element
    # to circle (the "what was tested" marker) on the screenshot.
    descriptions: dict[str, str] = {}
    circle_targets: dict[str, list] = {}
    steps: dict[str, list] = {}
    for check in summary.get('checks') or []:
        fid = check.get('id') or ''
        if fid and fid not in descriptions:
            descriptions[fid] = extract_flow_description(flows_dir, fid)
            circle_targets[fid] = derive_circle_targets(flows_dir, fid)
            steps[fid] = flow_to_steps(flows_dir, fid)

    build_report(summary, artifacts_dir, output, args.title, args.subtitle,
                 descriptions, circle_targets, steps)
    print(f'EXCEL_PATH={output}')
    return 0


if __name__ == '__main__':
    sys.exit(main())
