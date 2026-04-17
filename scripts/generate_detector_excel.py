from __future__ import annotations

import argparse
import json
import math
import re
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

PALETTE = {
    'navy': '0F172A',
    'slate': '334155',
    'muted': '64748B',
    'border': 'CBD5E1',
    'pass': '0F766E',
    'pass_bg': 'CCFBF1',
    'fail': 'B91C1C',
    'fail_bg': 'FEE2E2',
    'skip': '92400E',
    'skip_bg': 'FEF3C7',
}

BOUNDS_RE = re.compile(r'\[(\d+),(\d+)\]\[(\d+),(\d+)\]')
CROP_PAD = 30  # pixels of padding around the element when cropping


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description='Generate UI Change Detector Mobile Excel report from scan results.',
    )
    p.add_argument('--scan-json', required=True, help='Path to scan_results JSON file')
    p.add_argument('--output', required=True, help='Path for the output .xlsx file')
    p.add_argument('--screenshots-dir', default='', help='Directory containing screen PNGs')
    p.add_argument('--title', default='UI Change Detector Mobile')
    p.add_argument('--subtitle', default='')
    return p.parse_args()


# ===================================================================
# PIL: crop + annotate helpers (matching webapp-ui-detector pattern)
# ===================================================================

def parse_bounds(bounds_str: str) -> tuple[int, int, int, int] | None:
    """Parse '[x1,y1][x2,y2]' into (x1, y1, x2, y2)."""
    m = BOUNDS_RE.match(bounds_str)
    if not m:
        return None
    return int(m.group(1)), int(m.group(2)), int(m.group(3)), int(m.group(4))


def _load_font(size: int = 16):
    for path in (
        'arial.ttf',
        'arialbd.ttf',
        '/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf',
        '/usr/share/fonts/truetype/liberation/LiberationSans-Bold.ttf',
    ):
        try:
            return ImageFont.truetype(path, size)
        except (OSError, IOError):
            continue
    return ImageFont.load_default()


def crop_and_annotate(
    img: PILImage.Image,
    x1: int, y1: int, x2: int, y2: int,
    label: int,
    tmp_dir: Path,
    tag: str,
) -> Path:
    """
    Crop a region around (x1,y1)-(x2,y2) with CROP_PAD padding,
    draw a red circle + numbered label (matching webapp annotate.mjs style),
    and save to tmp_dir.
    """
    # Crop region with padding, clamped to image bounds
    cx1 = max(0, min(x1 - CROP_PAD, img.width - 1))
    cy1 = max(0, min(y1 - CROP_PAD, img.height - 1))
    cx2 = min(img.width, max(x2 + CROP_PAD, cx1 + 1))
    cy2 = min(img.height, max(y2 + CROP_PAD, cy1 + 1))

    if cx2 <= cx1 or cy2 <= cy1:
        return None

    crop = img.crop((cx1, cy1, cx2, cy2)).convert('RGBA')
    draw = ImageDraw.Draw(crop)

    # Element bounds relative to crop
    rx1 = x1 - cx1
    ry1 = y1 - cy1
    rx2 = x2 - cx1
    ry2 = y2 - cy1

    # Circle: centred on the element, radius = max(w,h)/2 + 14
    elem_w = rx2 - rx1
    elem_h = ry2 - ry1
    center_x = rx1 + elem_w / 2
    center_y = ry1 + elem_h / 2
    radius = max(elem_w, elem_h) / 2 + 14

    draw.ellipse(
        [center_x - radius, center_y - radius,
         center_x + radius, center_y + radius],
        outline=(220, 38, 38, 255),
        width=4,
    )

    # Numbered label pill (top-right of circle)
    font = _load_font(16)
    label_text = str(label)
    bbox = font.getbbox(label_text)
    lw = bbox[2] - bbox[0]
    lh = bbox[3] - bbox[1]
    pill_r = 14
    pill_cx = center_x + radius - 4
    pill_cy = center_y - radius + 4
    draw.ellipse(
        [pill_cx - pill_r, pill_cy - pill_r,
         pill_cx + pill_r, pill_cy + pill_r],
        fill=(220, 38, 38, 230),
    )
    draw.text(
        (pill_cx - lw / 2, pill_cy - lh / 2 - 2),
        label_text,
        font=font,
        fill=(255, 255, 255, 255),
    )

    out = tmp_dir / f'{tag}.png'
    crop.convert('RGB').save(str(out), 'PNG')
    return out


def annotate_full_screenshot(
    img: PILImage.Image,
    elements: list[dict],
    tmp_dir: Path,
    screen_id: str,
) -> Path | None:
    """Draw red circles + numbered labels on a full screenshot for all new elements."""
    overlay = img.copy().convert('RGBA')
    draw = ImageDraw.Draw(overlay)
    font = _load_font(16)
    drawn = 0

    for idx, elem in enumerate(elements, start=1):
        bounds = parse_bounds(elem.get('bounds', ''))
        if bounds is None:
            continue
        x1, y1, x2, y2 = bounds

        elem_w = x2 - x1
        elem_h = y2 - y1
        cx = x1 + elem_w / 2
        cy = y1 + elem_h / 2
        r = max(elem_w, elem_h) / 2 + 14

        draw.ellipse(
            [cx - r, cy - r, cx + r, cy + r],
            outline=(220, 38, 38, 255),
            width=4,
        )

        # Number label
        label = str(idx)
        bbox = font.getbbox(label)
        lw = bbox[2] - bbox[0]
        lh = bbox[3] - bbox[1]
        pill_r = 14
        pill_cx = cx + r - 4
        pill_cy = cy - r + 4
        draw.ellipse(
            [pill_cx - pill_r, pill_cy - pill_r,
             pill_cx + pill_r, pill_cy + pill_r],
            fill=(220, 38, 38, 230),
        )
        draw.text(
            (pill_cx - lw / 2, pill_cy - lh / 2 - 2),
            label,
            font=font,
            fill=(255, 255, 255, 255),
        )
        drawn += 1

    if drawn == 0:
        return None

    out = tmp_dir / f'{screen_id}_full_annotated.png'
    overlay.convert('RGB').save(str(out), 'PNG')
    return out


def _make_missing_thumbnail(
    img: PILImage.Image,
    label_text: str,
    idx: int,
    tmp_dir: Path,
    tag: str,
) -> Path | None:
    """Scaled-down screenshot with amber MISSING banner showing what element is absent."""
    try:
        thumb = img.copy().convert('RGBA')
        thumb.thumbnail((240, 400))
        draw = ImageDraw.Draw(thumb)
        font = _load_font(14)

        # Amber banner at top
        draw.rectangle([0, 0, thumb.width, 28], fill=(245, 158, 11, 220))
        banner = f'#{idx} MISSING'
        bbox = font.getbbox(banner)
        tw = bbox[2] - bbox[0]
        draw.text(((thumb.width - tw) / 2, 4), banner, font=font, fill=(255, 255, 255, 255))

        # Element name at bottom
        if label_text:
            small_font = _load_font(11)
            draw.rectangle([0, thumb.height - 22, thumb.width, thumb.height], fill=(0, 0, 0, 160))
            draw.text((4, thumb.height - 20), label_text[:40], font=small_font, fill=(255, 255, 255, 255))

        out = tmp_dir / f'{tag}.png'
        thumb.convert('RGB').save(str(out), 'PNG')
        return out
    except Exception:
        return None


def find_screenshot(screenshots_dir: Path, screen_id: str) -> Path | None:
    for name in (f'{screen_id}.png', f'{screen_id}_changes.png', f'{screen_id}_new_elements.png'):
        p = screenshots_dir / name
        if p.exists():
            return p
    return None


# ===================================================================
# openpyxl helpers
# ===================================================================

def style_header(cell, fill='navy'):
    cell.fill = PatternFill('solid', fgColor=PALETTE[fill])
    cell.font = Font(name='Aptos', bold=True, size=11, color='FFFFFF')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def card(ws, cell_range, title, value, fill_color, value_color='FFFFFF'):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(':')[0]]
    cell.value = f'{title}\n{value}'
    cell.fill = PatternFill('solid', fgColor=fill_color)
    cell.font = Font(name='Aptos', bold=True, size=17, color=value_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color=PALETTE['border'])
    for row in ws[cell_range]:
        for item in row:
            item.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def add_image_scaled(ws, path_str, anchor_cell, max_w=240, max_h=140):
    """Embed an image scaled to fit max_w x max_h (same as webapp)."""
    p = Path(path_str)
    if not p.exists():
        return
    try:
        img = XLImage(str(p))
    except Exception:
        return
    w, h = img.width, img.height
    ratio_w = max_w / max(w, 1)
    ratio_h = max_h / max(h, 1)
    ratio = min(ratio_w, ratio_h, 1)
    img.width = int(w * ratio)
    img.height = int(h * ratio)
    img.anchor = anchor_cell
    ws.add_image(img)


def load_scan(path: Path) -> dict:
    return json.loads(path.read_text(encoding='utf-8'))


# ===================================================================
# Sheet: Summary
# ===================================================================

def build_summary_sheet(ws, data: dict, title: str, subtitle: str):
    screens = data.get('screens', {})
    summary = data.get('summary', {})
    scan_ts = data.get('scan_timestamp', '')

    screens_scanned = summary.get('screens_scanned', len(screens))
    total_new = summary.get('total_new_elements', 0)
    passed = sum(1 for s in screens.values()
                 if s.get('new_element_count', 0) == 0 and s.get('removed_element_count', 0) == 0)
    failed = screens_scanned - passed
    total_removed = sum(s.get('removed_element_count', 0) for s in screens.values())

    # Title
    ws.merge_cells('A1:H1')
    ws['A1'].value = title
    ws['A1'].fill = PatternFill('solid', fgColor=PALETTE['navy'])
    ws['A1'].font = Font(name='Aptos Display', bold=True, size=22, color='FFFFFF')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 42

    # Subtitle
    sub_text = subtitle if subtitle else f'Scan: {scan_ts}'
    ws.merge_cells('A2:H2')
    ws['A2'].value = sub_text
    ws['A2'].fill = PatternFill('solid', fgColor=PALETTE['slate'])
    ws['A2'].font = Font(name='Aptos', size=12, color='FFFFFF')
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 28

    ws.row_dimensions[3].height = 8

    # Metric cards
    ws.row_dimensions[4].height = 52
    ws.row_dimensions[5].height = 52
    card(ws, 'A4:B5', 'Screens Scanned', screens_scanned, PALETTE['slate'])
    card(ws, 'C4:D5', 'Passed', passed, PALETTE['pass'])
    card(ws, 'E4:F5', 'Failed', failed, PALETTE['fail'])
    card(ws, 'G4:H5', 'Changes', f'{total_new} new / {total_removed} removed', PALETTE['skip'], value_color='000000')

    ws.row_dimensions[6].height = 8
    ws.row_dimensions[7].height = 8

    # Screen table
    headers = ['Screen ID', 'Status', 'New Elements', 'Removed Elements', 'Details']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=8, column=i, value=h)
        style_header(c)
    ws.row_dimensions[8].height = 24

    row = 9
    for screen_id in sorted(screens.keys()):
        screen = screens[screen_id]
        new_count = screen.get('new_element_count', 0)
        removed_count = screen.get('removed_element_count', 0)
        status = 'PASS' if (new_count == 0 and removed_count == 0) else 'FAIL'

        details_parts = []
        if new_count > 0:
            parts = []
            for elem in screen.get('new_elements', []):
                label = elem.get('text') or elem.get('content_desc') or elem.get('resource_id') or elem.get('class', '')
                parts.append(f'+{label}')
            details_parts.extend(parts)
        if removed_count > 0:
            parts = []
            for elem in screen.get('removed_elements', []):
                label = elem.get('text') or elem.get('content_desc') or elem.get('resource_id') or elem.get('class', '')
                parts.append(f'-{label}')
            details_parts.extend(parts)
        details = '; '.join(details_parts) if details_parts else ''

        ws.cell(row=row, column=1, value=screen_id)

        sc = ws.cell(row=row, column=2, value=status)
        if status == 'PASS':
            sc.fill = PatternFill('solid', fgColor=PALETTE['pass_bg'])
            sc.font = Font(name='Aptos', bold=True, color=PALETTE['pass'])
        else:
            sc.fill = PatternFill('solid', fgColor=PALETTE['fail_bg'])
            sc.font = Font(name='Aptos', bold=True, color=PALETTE['fail'])
        sc.alignment = Alignment(horizontal='center')

        ws.cell(row=row, column=3, value=new_count)
        ws.cell(row=row, column=4, value=removed_count)
        ws.cell(row=row, column=5, value=details)
        row += 1

    set_col_widths(ws, [30, 12, 14, 16, 60])

    ws.sheet_properties.tabColor = '3B82F6'
    ws.sheet_view.showGridLines = False
    ws.sheet_view.zoomScale = 90


# ===================================================================
# Sheet: per-screen (one sheet per screen with new elements)
# ===================================================================

MAX_SHEET_NAME = 31
SHEET_NAME_RE = re.compile(r'[\\/*?:\[\]]')


def safe_sheet_name(name: str, existing: set) -> str:
    base = SHEET_NAME_RE.sub('_', name).strip()[:MAX_SHEET_NAME] or 'Screen'
    candidate = base
    idx = 2
    while candidate in existing:
        suffix = f' ({idx})'
        candidate = (base[:MAX_SHEET_NAME - len(suffix)] + suffix)
        idx += 1
    existing.add(candidate)
    return candidate


def write_screen_sheet(
    ws,
    screen_id: str,
    screen_data: dict,
    screenshots_dir: Path | None,
    tmp_dir: Path,
):
    """Write a per-screen sheet with element rows + crop screenshots + full annotated image."""
    new_elements = screen_data.get('new_elements', [])
    removed_elements = screen_data.get('removed_elements', [])

    # Header
    ws['A1'] = f'Screen: {screen_id}'
    ws['A1'].font = Font(name='Aptos', bold=True, size=14, color=PALETTE['navy'])
    ws['A2'] = f'{len(new_elements)} new element(s), {len(removed_elements)} removed element(s) detected'
    ws['A2'].font = Font(name='Aptos', italic=True, size=10, color=PALETTE['muted'])

    # Table headers (row 4) — matching webapp layout
    headers = ['#', 'Change Type', 'Class', 'Text / Label', 'Resource ID', 'Bounds', 'Print-screen']
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=i, value=h)
        style_header(c)

    # Load screenshot for cropping
    screenshot_img = None
    if HAS_PIL and screenshots_dir and screenshots_dir.is_dir():
        ss_path = find_screenshot(screenshots_dir, screen_id)
        if ss_path:
            screenshot_img = PILImage.open(ss_path)

    # Element rows
    row = 5
    for idx, elem in enumerate(new_elements, start=1):
        ws.cell(row=row, column=1, value=idx)

        ct = ws.cell(row=row, column=2, value='New Element')
        ct.fill = PatternFill('solid', fgColor=PALETTE['fail_bg'])
        ct.font = Font(name='Aptos', bold=True, color=PALETTE['fail'])

        ws.cell(row=row, column=3, value=elem.get('class', ''))
        text = elem.get('text') or elem.get('content_desc') or ''
        ws.cell(row=row, column=4, value=text[:200])
        ws.cell(row=row, column=5, value=elem.get('resource_id', ''))
        ws.cell(row=row, column=6, value=elem.get('bounds', ''))

        ws.row_dimensions[row].height = 110

        # Crop + annotate this element and embed in Print-screen column
        bounds = parse_bounds(elem.get('bounds', ''))
        if screenshot_img and bounds:
            x1, y1, x2, y2 = bounds
            crop_path = crop_and_annotate(
                screenshot_img, x1, y1, x2, y2,
                label=idx,
                tmp_dir=tmp_dir,
                tag=f'{screen_id}_crop_{idx}',
            )
            add_image_scaled(ws, str(crop_path), f'G{row}')

        row += 1

    # Removed elements
    for idx, elem in enumerate(removed_elements, start=len(new_elements) + 1):
        ws.cell(row=row, column=1, value=idx)

        ct = ws.cell(row=row, column=2, value='Missing')
        ct.fill = PatternFill('solid', fgColor=PALETTE['skip_bg'])
        ct.font = Font(name='Aptos', bold=True, color=PALETTE['skip'])

        ws.cell(row=row, column=3, value=elem.get('class', ''))
        text = elem.get('text') or elem.get('content_desc') or ''
        ws.cell(row=row, column=4, value=text[:200])
        ws.cell(row=row, column=5, value=elem.get('resource_id', ''))
        bounds_str = elem.get('bounds', '')
        ws.cell(row=row, column=6, value=bounds_str)

        ws.row_dimensions[row].height = 110

        # If the removed element has bounds, crop and annotate
        bounds = parse_bounds(bounds_str)
        if screenshot_img and bounds:
            x1, y1, x2, y2 = bounds
            crop_path = crop_and_annotate(
                screenshot_img, x1, y1, x2, y2,
                label=idx,
                tmp_dir=tmp_dir,
                tag=f'{screen_id}_missing_{idx}',
            )
            if crop_path:
                add_image_scaled(ws, str(crop_path), f'G{row}')
        elif screenshot_img and HAS_PIL:
            # No bounds — show scaled screenshot with amber MISSING banner
            thumb = _make_missing_thumbnail(screenshot_img, text, idx, tmp_dir, f'{screen_id}_missing_thumb_{idx}')
            if thumb:
                add_image_scaled(ws, str(thumb), f'G{row}')

        row += 1

    # Full annotated screenshot at the bottom
    all_changes = new_elements + removed_elements
    if screenshot_img and all_changes:
        row += 2
        ws.cell(row=row, column=1, value='Full screenshot (with changes circled):').font = Font(bold=True)
        full_ann = annotate_full_screenshot(screenshot_img, all_changes, tmp_dir, screen_id)
        if full_ann:
            add_image_scaled(ws, str(full_ann), f'A{row + 1}', max_w=520, max_h=1000)

    set_col_widths(ws, [6, 14, 28, 36, 36, 22, 40])


# ===================================================================
# Main
# ===================================================================

def main():
    args = parse_args()
    scan_path = Path(args.scan_json)
    output_path = Path(args.output)
    screenshots_dir = Path(args.screenshots_dir) if args.screenshots_dir else None

    if not HAS_PIL:
        print('WARNING: Pillow not installed — screenshots will be skipped', file=sys.stderr)

    data = load_scan(scan_path)
    screens = data.get('screens', {})

    wb = Workbook()

    with tempfile.TemporaryDirectory(prefix='detector_excel_') as tmp_dir_str:
        tmp_dir = Path(tmp_dir_str)

        # Sheet 1: Summary
        ws_summary = wb.active
        ws_summary.title = 'Summary'
        build_summary_sheet(ws_summary, data, args.title, args.subtitle)

        # Per-screen sheets (only for screens with new elements)
        existing_names = {'Summary'}
        for screen_id in sorted(screens.keys()):
            screen_data = screens[screen_id]
            if screen_data.get('new_element_count', 0) == 0 and screen_data.get('removed_element_count', 0) == 0:
                continue
            name = safe_sheet_name(screen_id, existing_names)
            ws = wb.create_sheet(title=name)
            write_screen_sheet(ws, screen_id, screen_data, screenshots_dir, tmp_dir)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(str(output_path))

    print(f'EXCEL_PATH={output_path}')
    sys.exit(0)


if __name__ == '__main__':
    main()
