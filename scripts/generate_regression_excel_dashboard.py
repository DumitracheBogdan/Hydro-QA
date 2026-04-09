from __future__ import annotations

import argparse
import json
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

PALETTE = {
    'navy': '0F172A',
    'slate': '334155',
    'muted': '64748B',
    'border': 'CBD5E1',
    'pass': '0F766E',
    'pass_bg': 'CCFBF1',
    'fail': 'B91C1C',
    'fail_bg': 'FEE2E2',
    'skip': '92400E',
    'skip_bg': 'FEF3C7',
}

DEFAULT_ROOT = Path(r'c:\work\QA TRacker')
DEFAULT_COMBINED_JSON = DEFAULT_ROOT / 'qa-artifacts' / 'infra-regression' / 'tmp-combined-184-doublecheck.json'
DEFAULT_OUTPUT = DEFAULT_ROOT / 'qa-artifacts' / 'infra-regression' / 'Hydrocert_DEV_Regression_Dashboard_2026-03-06.xlsx'


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description='Generate Hydrocert regression Excel dashboard.')
    parser.add_argument('--combined-json', default=str(DEFAULT_COMBINED_JSON))
    parser.add_argument('--output', default=str(DEFAULT_OUTPUT))
    parser.add_argument('--title', default='Hydrocert DEV Regression Report')
    parser.add_argument('--subtitle', default='Generated from combined run summary')
    return parser.parse_args()


def autosize(ws, min_width=12, max_width=58):
    for col_cells in ws.columns:
        col_letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for cell in col_cells:
            value = '' if cell.value is None else str(cell.value)
            if '\n' in value:
                value = max(value.splitlines(), key=len)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, max_len + 2))


def style_header(cell, fill='navy', size=11, color='FFFFFF'):
    cell.fill = PatternFill('solid', fgColor=PALETTE[fill])
    cell.font = Font(name='Aptos', bold=True, size=size, color=color)
    cell.alignment = Alignment(horizontal='center', vertical='center')


def style_table_header_row(ws, row_idx, fill='navy'):
    for cell in ws[row_idx]:
        style_header(cell, fill=fill)


def card(ws, cell_range, title, value, fill_color, value_color='FFFFFF'):
    ws.merge_cells(cell_range)
    cell = ws[cell_range.split(':')[0]]
    cell.value = f'{title}\n{value}'
    cell.fill = PatternFill('solid', fgColor=fill_color)
    cell.font = Font(name='Aptos', bold=True, size=17, color=value_color)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color=PALETTE['border'])
    for row in ws[cell_range]:
        for item in row:
            item.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def add_table(ws, start_row, end_row, end_col, table_name):
    ref = f'A{start_row}:{get_column_letter(end_col)}{end_row}'
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name='TableStyleMedium2',
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def load_summary(path: Path) -> dict:
    return json.loads(path.read_text(encoding='utf-8'))


def unique_preserve_order(values):
    seen = set()
    result = []
    for value in values:
        if not value or value in seen:
            continue
        seen.add(value)
        result.append(value)
    return result


def extract_paths(*texts):
    import re

    matches = []
    pattern = re.compile(r'(/[a-zA-Z0-9][a-zA-Z0-9\-._~/%?=&:]*)(?=\s|$|,|\)|\|)')
    for text in texts:
        if not text:
            continue
        matches.extend(match.group(1) for match in pattern.finditer(str(text)))
    return unique_preserve_order(matches)


def safe_excel_value(value):
    if value is None:
        return ''
    if isinstance(value, str):
        return ILLEGAL_CHARACTERS_RE.sub('', value)
    return value


def localize_status(status: str) -> str:
    mapping = {
        'PASS': 'PASS',
        'FAIL': 'FAIL',
        'SKIP': 'SKIP',
    }
    return mapping.get(str(status or '').upper(), safe_excel_value(status))


def normalize_evidence_paths(check: dict) -> list[str]:
    evidence = check.get('evidence') or []
    result = []
    suite = str(check.get('suite', '')).strip()
    for item in unique_preserve_order(evidence):
        raw = str(item or '').strip()
        if not raw:
            continue
        normalized = raw.replace('\\', '/')
        filename = Path(normalized).name
        if '/screenshots/' in normalized and suite:
            result.append(f'suites/{suite}/screenshots/{filename}')
        elif '/downloads/' in normalized and suite:
            result.append(f'suites/{suite}/downloads/{filename}')
        else:
            result.append(filename or normalized)
    return result


def format_evidence(check: dict) -> str:
    entries = normalize_evidence_paths(check)
    if not entries:
        return ''

    formatted = []
    for entry in entries:
        lower = entry.lower()
        if lower.endswith(('.png', '.jpg', '.jpeg', '.webp')):
            formatted.append(f'Print-screen: {entry}')
        else:
            formatted.append(f'File: {entry}')
    return '\n'.join(formatted)


def build_issue_summary(check: dict) -> str:
    suite = str(check.get('suite', '')).upper()
    check_id = str(check.get('id', '')).upper()
    details = str(check.get('details', '') or '').strip()
    test = str(check.get('test', '') or '').strip()

    mapping = {
        ('ROLE06', 'RA01'): 'Swagger is public on dev and can be accessed without authentication.',
        ('ROLE06', 'RA02'): 'A user with normal role can create or delete customers.',
        ('ROLE06', 'RA03'): 'A user sees visits that don\'t belong to them.',
        ('ROLE06', 'RA04'): 'A user sees other people\'s absences.',
        ('ROLE06', 'RA05'): 'A user can view activity logs.',
        ('ROLE06', 'RA06'): 'A user can read global reference data.',
        ('ROLERO06', 'RR01'): 'Swagger is public on prod and can be accessed without authentication.',
        ('ROLERO06', 'RR02'): 'QA user account did not authenticate correctly on prod.',
        ('ROLERO06', 'RR03'): 'A user sees visits that don\'t belong to them on prod.',
        ('ROLERO06', 'RR04'): 'A user sees other people\'s absences on prod.',
        ('ROLERO06', 'RR05'): 'A user can view activity logs on prod.',
        ('ROLERO06', 'RR06'): 'A user can read global reference data on prod.',
        ('ESS25', 'E04'): 'Main JS bundle does not have a cache header.',
        ('ESS25', 'E05'): '/health endpoint does not respond in the expected format.',
        ('ESS25', 'E06'): '/health payload does not match the contract verified by the test.',
        ('NEW60', 'R07'): 'Missing Strict-Transport-Security header.',
        ('NEW60', 'R08'): 'Missing X-Content-Type-Options: nosniff header.',
        ('NEW60', 'R09'): 'Missing anti-frame protection: X-Frame-Options or CSP frame-ancestors.',
        ('NEW60', 'R10'): 'Web app allows TRACE method and should block it.',
        ('API34', 'L08'): 'Mixed API burst exceeds the set performance threshold.',
        ('NEW60', 'R43'): 'Mixed API burst exceeds the senior performance threshold.',
        ('DEEP32', 'P02'): '/users/profile/me endpoint responds too slowly vs threshold.',
        ('DEEP32', 'P04'): '/visits/calendar-filter has latency spike above threshold.',
        ('DEEP32', 'U03'): 'Visits List loaded without any rows.',
        ('DEEP32', 'U04'): 'Could not open Visit Details from first row.',
        ('DEEP32', 'U05'): 'Could not validate the tab rail in Visit Details.',
        ('DEEP32', 'U11'): '5xx responses appeared during app traversal.',
        ('DEEP32', 'U12'): 'Console errors appeared during app traversal.',
        ('UI22', 'U11'): 'Eye button in Planner did not correctly open the edit page.',
        ('UI22', 'U12'): 'Map on the edit page did not remain stable on refresh.',
    }

    summary = mapping.get((suite, check_id))
    if summary:
        if details:
            return f'{summary} Details: {details}.'
        return summary

    if '5xx' in details.lower():
        return f'Application returned 5xx responses in this flow. Details: {details}.'
    if 'console' in test.lower() or 'console' in details.lower():
        return f'Console errors appeared during the run. Details: {details}.'
    if 'cache-control' in details.lower():
        return f'Tested resource does not have Cache-Control. Details: {details}.'
    if 'rows=0' in details.lower():
        return 'Page or table loaded without rows, although the test expected data.'
    if any(token in test.lower() for token in ['p95', 'latency', 'burst', 'load']) or any(token in details.lower() for token in ['p95', 'avg=', 'latency']):
        return f'Performance test exceeded the set threshold. Details: {details or test}.'
    if details:
        return f'Result does not match the test expectation. Details: {details}.'
    return f'Result does not match the test expectation: {test}.'


def manual_target_hint(check: dict) -> str:
    suite = str(check.get('suite', '')).upper()
    check_id = str(check.get('id', '')).upper()

    if suite == 'ROLE06':
        mapping = {
            'RA01': '/api, /api-json, /auth/register',
            'RA02': '/customers -> POST /customers, DELETE /customers/{id}',
            'RA03': '/visits/calendar-filter',
            'RA04': '/users/absences',
            'RA05': '/activity-logs',
            'RA06': '/sites, /products, /job-types, /skills, /contracts',
        }
        return mapping.get(check_id, '')

    if suite == 'API34':
        mapping = {
            'I01': 'GET /',
            'I02': 'GET /health',
            'I03': '/dashboard, /customers, /visits-list, /planner, /visits, /visits/addnewvisit',
            'I04': 'Web host DNS',
            'I05': 'API host DNS',
            'I06': 'Web TLS certificate',
            'I07': 'API TLS certificate',
            'A01': 'POST /auth/login',
            'A02': 'POST /auth/login',
            'A03': 'GET /users/profile/me',
            'A04': 'GET /users/profile/me',
            'A05': 'GET /users/profile/me',
            'A06': 'GET /users',
            'A07': 'GET /users',
            'A08': 'GET /customers/filtered?page=1&limit=20',
            'A09': 'GET /customers/filtered?page=1&limit=1',
            'A10': 'GET /customers/filtered?page=1&limit=5',
            'A11': 'GET /customers/filtered?page=1..2&limit=20',
            'A12': 'GET /visits/calendar-filter?...&limit=50',
            'A13': 'GET /visits/calendar-filter?...&limit=10',
            'A14': 'GET /visits/calendar-filter?...&limit=50',
            'A15': 'GET /products',
            'A16': 'GET /sample-types',
            'A17': 'GET /labs',
            'A18': 'GET /job-types',
            'A19': 'GET /users/absences',
            'L01': 'GET /health',
            'L02': 'GET /users/profile/me',
            'L03': 'GET /customers/filtered?page=1&limit=20',
            'L04': 'GET /visits/calendar-filter?...&limit=50',
            'L05': 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20 + /visits/calendar-filter?...&limit=50',
            'L06': 'GET /customers/filtered?page=1&limit=20',
            'L07': 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20 + /visits/calendar-filter?...&limit=50',
            'L08': 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20 + /visits/calendar-filter?...&limit=50',
        }
        return mapping.get(check_id, '')

    if suite == 'UI22':
        mapping = {
            'U01': '/dashboard -> /login',
            'U02': '/dashboard',
            'U03': '/customers',
            'U04': '/customers',
            'U05': '/visits-list',
            'U06': '/visits-list',
            'U07': '/visits-list -> first row -> /visits/details/{id}',
            'U08': '/visits-list -> first row -> /visits/details/{id} -> Attachments / Visit Details tabs',
            'U09': '/visits-list -> first row -> /visits/details/{id} -> Attachments tab',
            'U10': '/planner',
            'U11': '/planner -> Events View -> eye action -> /visits/edit/{id}',
            'U12': '/planner -> Events View -> eye action -> /visits/edit/{id}',
            'U13': '/visits/addnewvisit',
            'U14': '/visits/addnewvisit',
            'U15': '/dashboard, /customers, /visits-list, /planner, /visits/details/{id}, /visits/edit/{id}, /visits/addnewvisit',
            'U16': '/dashboard, /customers, /visits-list, /planner, /visits/details/{id}, /visits/edit/{id}, /visits/addnewvisit',
            'U17': '/dashboard, /customers, /visits-list, /planner, /visits/details/{id}, /visits/edit/{id}, /visits/addnewvisit',
        }
        return mapping.get(check_id, '')

    if suite == 'DEEP32':
        mapping = {
            'I01': 'GET /',
            'I02': '/dashboard, /customers, /visits-list, /planner, /visits/addnewvisit, /visits',
            'I03': '/dashboard -> /login',
            'I04': '/dashboard',
            'I05': 'POST /auth/login',
            'I06': 'POST /auth/login',
            'A01': 'GET /health',
            'A02': 'GET /users/profile/me',
            'A03': 'GET /users',
            'A04': 'GET /customers/filtered?page=1&limit=20',
            'A05': 'GET /visits/calendar-filter?...&limit=50',
            'A06': 'GET /job-types',
            'A07': 'GET /products',
            'A08': 'GET /sample-types',
            'A09': 'GET /labs',
            'A10': 'GET /users/absences',
            'U01': '/customers',
            'U02': '/customers',
            'U03': '/visits-list',
            'U04': '/visits-list -> first row -> /visits/details/{id}',
            'U05': '/visits-list -> first row -> /visits/details/{id} -> Attachments / Visit Details tabs',
            'U06': '/planner',
            'U07': '/planner -> Events View -> eye action -> /visits/edit/{id}',
            'U08': '/planner -> Events View -> eye action -> /visits/edit/{id}',
            'U09': '/visits/addnewvisit',
            'U10': '/dashboard, /customers, /visits-list, /planner, /visits/details/{id}, /visits/edit/{id}, /visits/addnewvisit',
            'U11': '/dashboard, /customers, /visits-list, /planner, /visits/details/{id}, /visits/edit/{id}, /visits/addnewvisit',
            'U12': '/dashboard, /customers, /visits-list, /planner, /visits/details/{id}, /visits/edit/{id}, /visits/addnewvisit',
            'P01': 'GET /health',
            'P02': 'GET /users/profile/me',
            'P03': 'GET /customers/filtered?page=1&limit=20',
            'P04': 'GET /visits/calendar-filter?...&limit=50',
        }
        return mapping.get(check_id, '')

    if suite == 'ESS25':
        mapping = {
            'E01': 'GET http://web root',
            'E02': 'GET http://api root or /health',
            'E03': 'GET /',
            'E04': 'GET / -> main JS bundle',
            'E05': 'GET /health',
            'E06': 'GET /health',
            'E07': 'POST /auth/login',
            'E08': 'GET /users/profile/me',
            'E09': 'POST /auth/refresh, /auth/refresh-token, /auth/refreshToken',
            'E10': 'POST /auth/login',
            'E11': 'POST /auth/login -> JWT claims',
            'E12': 'GET /users',
            'E13': 'GET /customers/filtered?page=1&limit=50',
            'E14': 'GET /visits/calendar-filter?...&limit=50',
            'E15': 'GET /visits/calendar-filter?...&limit=50',
            'E16': 'GET /visits/calendar-filter?...&limit=50 + GET /users',
            'E17': 'GET /visits/calendar-filter?...&limit=50 + GET /users',
            'E18': 'GET /customers/filtered?page=1..2&limit=50',
            'E19': 'GET /customers/filtered?page=1&limit=20',
            'E20': 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20',
            'E21': 'GET /users/profile/me',
            'E22': '/dashboard, /customers, /visits-list, /planner',
            'E23': '/dashboard -> logout',
            'E24': '/dashboard after logout',
            'E25': 'Desktop telemetry for the essential UI run',
        }
        return mapping.get(check_id, '')

    if suite == 'NEW60':
        mapping = {
            'R01': 'GET /',
            'R02': 'GET /health',
            'R03': 'Web host DNS',
            'R04': 'API host DNS',
            'R05': 'Web TLS certificate',
            'R06': 'API TLS certificate',
            'R07': 'GET /',
            'R08': 'GET /',
            'R09': 'GET /',
            'R10': 'TRACE / and TRACE /health',
            'R11': 'POST /auth/login',
            'R12': 'POST /auth/login',
            'R13': 'POST /auth/login -> JWT',
            'R14': 'POST /auth/login -> JWT',
            'R15': 'POST /auth/login -> JWT',
            'R16': 'POST /auth/login',
            'R17': 'POST /auth/login',
            'R18': 'POST /auth/login',
            'R19': 'GET /users/profile/me',
            'R20': 'GET /users/profile/me',
            'R21': 'GET /users',
            'R22': 'GET /users',
            'R23': 'GET /users?unknownParam=1',
            'R24': 'GET /customers/filtered?page=-1&limit=20',
            'R25': 'GET /customers/filtered?page=1&limit=0',
            'R26': 'GET /customers/filtered?page=1&limit=5000',
            'R27': 'GET /customers/filtered?page=1&limit=20&search=<script>',
            'R28': 'GET /customers/filtered?page=1&limit=20&search=SQL-like payload',
            'R29': 'GET /visits/calendar-filter?startDate=invalid&endDate=invalid&page=1&limit=20',
            'R30': 'GET /visits/calendar-filter?startDate> endDate',
            'R31': 'GET /visits/calendar-filter?...&limit=10',
            'R32': 'GET /users/absences?startDate=invalid&endDate=invalid',
            'R33': 'GET /users/absences?startDate> endDate',
            'R34': 'GET /products?foo=bar&baz=1',
            'R35': 'GET /definitely-not-existing-endpoint-xyz',
            'R36': 'GET /health',
            'R37': 'GET /users/profile/me',
            'R38': 'GET /customers/filtered?page=1&limit=20',
            'R39': 'GET /visits/calendar-filter?...&limit=50',
            'R40': 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20 + /visits/calendar-filter?...&limit=50',
            'R41': 'GET /customers/filtered?page=1&limit=20',
            'R42': 'GET /users',
            'R43': 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20 + /visits/calendar-filter?...&limit=50',
            'R44': 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20 + /visits/calendar-filter?...&limit=50',
            'R45': 'GET /health',
            'R46': '/dashboard -> /login',
            'R47': '/planner',
            'R48': '/this-route-should-not-exist-senior-check',
            'R49': '/dashboard -> /customers -> /planner -> browser back/forward',
            'R50': '/visits-list -> first row -> /visits/details/{id}',
            'R51': '/planner -> Events View -> eye action -> /visits/edit/{id}',
            'R52': '/visits/edit/{id}',
            'R53': '/dashboard -> sessionStorage',
            'R54': '/dashboard -> localStorage',
            'R55': 'Fresh browser context -> /dashboard',
        }
        return mapping.get(check_id, '')

    if suite == 'SOAK11':
        if check_id == 'SOAK-00':
            return 'POST /auth/login'
        return 'GET /health + /users/profile/me + /customers/filtered?page=1&limit=20 + /visits/calendar-filter?...&limit=20 + /products'

    return ''


def target_hint(check: dict) -> str:
    manual = manual_target_hint(check)
    if manual:
        return manual

    text = ' '.join([
        str(check.get('suite', '')),
        str(check.get('area', '')),
        str(check.get('test', '')),
        str(check.get('details', '')),
    ]).lower()
    paths = extract_paths(check.get('test', ''), check.get('details', ''))
    if paths:
        return ', '.join(paths[:3])

    hints = [
        ('/api and /api-json', ['swagger', 'openapi', 'docs']),
        ('/auth/register', ['self-registration', 'register']),
        ('/auth/login', ['login']),
        ('/users/profile/me', ['profile']),
        ('/users', ['users endpoint']),
        ('/customers/filtered?page=1&limit=20', ['customers filtered', 'customers page', 'customer']),
        ('/customers', ['create customer', 'delete customer']),
        ('/visits/calendar-filter', ['calendar filter', 'visits filter', 'own visits', 'visit']),
        ('/users/absences', ['absences']),
        ('/activity-logs', ['activity logs']),
        ('/products', ['products']),
        ('/sample-types', ['sample types']),
        ('/labs', ['labs']),
        ('/job-types', ['job types']),
        ('/skills', ['skills']),
        ('/sites', ['sites']),
        ('/contracts', ['contracts']),
        ('/health', ['health']),
        ('/dashboard', ['dashboard']),
        ('/planner', ['planner']),
        ('/visits-list', ['visits list']),
        ('/visits/addnewvisit', ['add new visit']),
        ('/visits', ['visits']),
        ('/settings', ['settings']),
        ('/team-management', ['team management']),
        ('/schedule', ['schedule']),
        ('/', ['web root', 'root']),
    ]
    for endpoint, keywords in hints:
        if any(keyword in text for keyword in keywords):
            return endpoint
    return ''


def build_steps(check: dict) -> str:
    existing = check.get('steps')
    if isinstance(existing, list) and existing:
        return '\n'.join(f'{idx}. {step}' for idx, step in enumerate(existing, start=1))
    if isinstance(existing, str) and existing.strip():
        return existing.strip()

    suite = str(check.get('suite', '')).lower()
    area = str(check.get('area', '')).lower()
    test = str(check.get('test', ''))
    test_lc = test.lower()
    endpoint = target_hint(check)

    def lines(*items):
        return '\n'.join(f'{idx}. {item}' for idx, item in enumerate(items, start=1))

    if 'swagger docs and anonymous self-registration' in test_lc:
        return lines(
            'Open documentation endpoints without authentication.',
            'Access `/api` and `/api-json`, then try `POST /auth/register` without login.',
            'Verify that documentation or self-registration remain anonymously accessible.',
        )

    if 'create or delete customers' in test_lc:
        return lines(
            'Log in with an account with `user` role and keep the bearer token.',
            'Call `POST /customers` with a unique test name, then `DELETE /customers/{id}` if create succeeds.',
            'Verify that both actions are blocked for the `user` role.',
        )

    if 'only sees own visits' in test_lc:
        return lines(
            'Log in with an account with `user` role and keep the bearer token.',
            'Call `GET /visits/calendar-filter` on an interval that returns data.',
            'Verify the response contains only visits assigned to the logged-in user.',
        )

    if 'team absences' in test_lc:
        return lines(
            'Log in with an account with `user` role and keep the bearer token.',
            'Call `GET /users/absences` on an interval with known absences.',
            'Verify the response does not expose other employees\' absences.',
        )

    if 'activity logs' in test_lc:
        return lines(
            'Log in with an account with `user` role and keep the bearer token.',
            'Call `GET /activity-logs`.',
            'Verify access is blocked or limited according to expected permissions.',
        )

    if 'global reference data' in test_lc:
        return lines(
            'Log in with an account with `user` role and keep the bearer token.',
            'Call common reference data endpoints: `/sites`, `/products`, `/job-types`, `/skills`, `/contracts`.',
            'Verify the `user` role cannot read global reference data unless explicitly allowed.',
        )

    if any(keyword in test_lc for keyword in ['without token', 'anonymous']):
        return lines(
            f'Call `{endpoint or "the tested endpoint"}` without `Authorization` header.',
            'Check the status code and payload.',
            f'Verify that the result matches the test expectation: {test}.',
        )

    if any(keyword in test_lc for keyword in ['invalid token', 'tampered token']):
        return lines(
            'Authenticate first and obtain a valid bearer token.',
            f'Call `{endpoint or "the tested endpoint"}` with an invalid or tampered token.',
            f'Verify that the result matches the test expectation: {test}.',
        )

    if 'valid login' in test_lc or ('login' in test_lc and 'invalid' not in test_lc):
        return lines(
            'Open the API client used for verification.',
            'Call `POST /auth/login` with the QA credentials used by the pipeline.',
            f'Verify that the result matches the test expectation: {test}.',
        )

    if 'invalid login' in test_lc:
        return lines(
            'Open the API client used for verification.',
            'Call `POST /auth/login` with the QA email and an invalid password.',
            f'Verify that the result matches the test expectation: {test}.',
        )

    if 'burst' in test_lc or 'p95' in test_lc or 'avg <=' in test_lc or 'avg <' in test_lc or 'load' in area or 'perf' in area:
        return lines(
            f'Authenticate if needed, then hit `{endpoint or "the tested endpoint"}`.',
            'Repeat the request with the sequence or concurrency described in the test.',
            'Measure latency or errors and compare the result with the threshold defined by the test.',
        )

    if any(keyword in suite for keyword in ['api', 'roleaccess']) or any(keyword in area for keyword in ['api', 'auth', 'access control', 'security']):
        auth_step = 'Authenticate with the QA account if the endpoint requires a bearer token.'
        if 'public' in test_lc or 'http' in test_lc or 'tls' in test_lc or 'dns' in test_lc:
            auth_step = 'No login needed if the endpoint is not protected.'
        target_step = f'Call `{endpoint or "the endpoint used by the test"}` with the same query or payload used by automation.'
        if ' + ' in endpoint or ', ' in endpoint:
            target_step = f'Use exactly these endpoints or flows as in automation: `{endpoint}`.'
        return lines(
            auth_step,
            target_step,
            f'Compare the response with the expected test result: {test}.',
        )

    if 'ui' in suite or 'web' in area:
        route = endpoint or 'the tested page'
        route_step = f'Log in if needed, then navigate to `{route}`.'
        if '->' in route or ', ' in route:
            route_step = f'Log in if needed, then follow this flow in the application: `{route}`.'
        return lines(
            'Open the web application on the same environment where the regression ran.',
            route_step,
            f'Execute the action described by the test and compare the actual result with the expectation: {test}.',
        )

    return lines(
        'Open the same environment where the regression ran.',
        'Repeat the action from the test name on the same page or endpoint used by automation.',
        f'Compare the actual result with the expected test behavior: {test}.',
    )


def build_workbook(summary: dict, output_path: Path, title: str, subtitle: str):
    checks = [
        {
            **check,
            'target_path': target_hint(check),
            'issue_summary': build_issue_summary(check),
            'steps_to_reproduce': build_steps(check),
            'evidence_text': format_evidence(check),
            'evidence_links': normalize_evidence_paths(check),
        }
        for check in summary['checks']
    ]
    totals = Counter(check['status'] for check in checks)
    suite_stats: dict[str, Counter] = defaultdict(Counter)

    for check in checks:
        suite_stats[check['suite']][check['status']] += 1
        suite_stats[check['suite']]['TOTAL'] += 1

    failed_rows = [check for check in checks if check['status'] == 'FAIL']
    suite_order = [suite['suite'] for suite in summary.get('suiteRuns', [])] or sorted(suite_stats.keys())

    wb = Workbook()
    wb.properties.creator = 'Codex'
    wb.properties.title = title
    wb.properties.subject = 'Hydrocert Regression Report'
    wb.properties.description = subtitle

    dashboard = wb.active
    dashboard.title = 'Summary'
    dashboard.sheet_view.showGridLines = False
    dashboard.sheet_properties.tabColor = '2563EB'

    dashboard['A1'] = title
    dashboard['A2'] = f"{subtitle} | Generated at {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    dashboard['A1'].font = Font(name='Aptos Display', bold=True, size=22, color='FFFFFF')
    dashboard['A2'].font = Font(name='Aptos', italic=True, size=10, color='E2E8F0')
    dashboard.merge_cells('A1:H1')
    dashboard.merge_cells('A2:H2')
    for row in ('A1:H1', 'A2:H2'):
        for cells in dashboard[row]:
            for cell in cells:
                cell.fill = PatternFill('solid', fgColor=PALETTE['navy'])

    card(dashboard, 'A4:B6', 'Total Tests', len(checks), PALETTE['slate'])
    card(dashboard, 'C4:D6', 'Passed', totals['PASS'], PALETTE['pass'])
    card(dashboard, 'E4:F6', 'Failed', totals['FAIL'], PALETTE['fail'])
    card(dashboard, 'G4:H6', 'Skipped', totals['SKIP'], PALETTE['skip'])

    dashboard['A8'] = 'Suite Summary'
    dashboard['A8'].font = Font(name='Aptos', bold=True, size=14, color=PALETTE['navy'])
    headers = ['Suite', 'Total', 'Passed', 'Failed', 'Skipped']
    for idx, header in enumerate(headers, start=1):
        dashboard.cell(9, idx, header)
    style_table_header_row(dashboard, 9)

    row = 10
    for suite in suite_order:
        stats = suite_stats[suite]
        dashboard.cell(row, 1, suite)
        dashboard.cell(row, 2, stats['TOTAL'])
        dashboard.cell(row, 3, stats['PASS'])
        dashboard.cell(row, 4, stats['FAIL'])
        dashboard.cell(row, 5, stats['SKIP'])
        row += 1

    dashboard['G8'] = 'Current Failures'
    dashboard['G8'].font = Font(name='Aptos', bold=True, size=14, color=PALETTE['navy'])
    failed_headers = ['Suite', 'ID', 'Area', 'Test']
    for idx, header in enumerate(failed_headers, start=7):
        dashboard.cell(9, idx, header)
    style_table_header_row(dashboard, 9)

    row = 10
    for check in failed_rows:
        dashboard.cell(row, 7, check['suite'])
        dashboard.cell(row, 8, check['id'])
        dashboard.cell(row, 9, check['area'])
        dashboard.cell(row, 10, check['test'])
        row += 1

    thin = Side(style='thin', color=PALETTE['border'])
    for row_cells in dashboard.iter_rows():
        for cell in row_cells:
            if cell.row >= 9:
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if cell.row != 9:
                    cell.alignment = Alignment(vertical='center', horizontal='left', wrap_text=True)

    dashboard.column_dimensions['A'].width = 18
    dashboard.column_dimensions['B'].width = 12
    dashboard.column_dimensions['C'].width = 12
    dashboard.column_dimensions['D'].width = 12
    dashboard.column_dimensions['E'].width = 12
    dashboard.column_dimensions['F'].width = 12
    dashboard.column_dimensions['G'].width = 18
    dashboard.column_dimensions['H'].width = 12
    dashboard.column_dimensions['I'].width = 16
    dashboard.column_dimensions['J'].width = 44

    all_tests = wb.create_sheet('All Tests')
    all_tests.freeze_panes = 'A2'
    all_tests.sheet_properties.tabColor = PALETTE['pass']
    all_headers = ['#', 'Suite', 'ID', 'Area', 'Status', 'Test', 'Path / Endpoint', 'Issue Summary', 'Technical Details', 'Reproduction Steps', 'Evidence']
    all_tests.append(all_headers)
    style_table_header_row(all_tests, 1)
    for idx, check in enumerate(checks, start=1):
        all_tests.append([
            idx,
            safe_excel_value(check['suite']),
            safe_excel_value(check['id']),
            safe_excel_value(check['area']),
            safe_excel_value(localize_status(check['status'])),
            safe_excel_value(check['test']),
            safe_excel_value(check.get('target_path', '')),
            safe_excel_value(check.get('issue_summary', '')),
            safe_excel_value(check.get('details', '')),
            safe_excel_value(check.get('steps_to_reproduce', '')),
            safe_excel_value(check.get('evidence_text', '')),
        ])
    add_table(all_tests, 1, len(checks) + 1, len(all_headers), 'AllTestsTable')
    autosize(all_tests, max_width=72)

    for row_idx in range(2, len(checks) + 2):
        status_cell = all_tests.cell(row_idx, 5)
        fill = PALETTE['pass_bg']
        font_color = PALETTE['pass']
        if status_cell.value == 'FAIL':
            fill = PALETTE['fail_bg']
            font_color = PALETTE['fail']
        elif status_cell.value == 'SKIP':
            fill = PALETTE['skip_bg']
            font_color = PALETTE['skip']
        status_cell.fill = PatternFill('solid', fgColor=fill)
        status_cell.font = Font(name='Aptos', bold=True, color=font_color)
        status_cell.alignment = Alignment(horizontal='center')
        all_tests.cell(row_idx, 6).alignment = Alignment(wrap_text=True, vertical='top')
        all_tests.cell(row_idx, 7).alignment = Alignment(wrap_text=True, vertical='top')
        all_tests.cell(row_idx, 8).alignment = Alignment(wrap_text=True, vertical='top')
        all_tests.cell(row_idx, 9).alignment = Alignment(wrap_text=True, vertical='top')
        all_tests.cell(row_idx, 10).alignment = Alignment(wrap_text=True, vertical='top')
        all_tests.cell(row_idx, 11).alignment = Alignment(wrap_text=True, vertical='top')
        evidence_cell = all_tests.cell(row_idx, 11)
        evidence_links = checks[row_idx - 2].get('evidence_links', [])
        if evidence_links:
            relative_target = evidence_links[0]
            if (output_path.parent / relative_target).exists():
                evidence_cell.hyperlink = relative_target
                evidence_cell.style = 'Hyperlink'

    failed_sheet = wb.create_sheet('Failed Details')
    failed_sheet.freeze_panes = 'A2'
    failed_sheet.sheet_properties.tabColor = PALETTE['fail']
    failed_detail_headers = ['Suite', 'ID', 'Area', 'Status', 'Test', 'Path / Endpoint', 'Issue Summary', 'Technical Details', 'Reproduction Steps', 'Evidence']
    failed_sheet.append(failed_detail_headers)
    style_table_header_row(failed_sheet, 1)
    for row_idx, check in enumerate(failed_rows, start=2):
        failed_sheet.append([
            safe_excel_value(check['suite']),
            safe_excel_value(check['id']),
            safe_excel_value(check['area']),
            safe_excel_value(localize_status(check['status'])),
            safe_excel_value(check['test']),
            safe_excel_value(check.get('target_path', '')),
            safe_excel_value(check.get('issue_summary', '')),
            safe_excel_value(check.get('details', '')),
            safe_excel_value(check.get('steps_to_reproduce', '')),
            safe_excel_value(check.get('evidence_text', '')),
        ])
        evidence_cell = failed_sheet.cell(row_idx, 10)
        evidence_links = check.get('evidence_links', [])
        if evidence_links:
            relative_target = evidence_links[0]
            if (output_path.parent / relative_target).exists():
                evidence_cell.hyperlink = relative_target
                evidence_cell.style = 'Hyperlink'
    add_table(failed_sheet, 1, max(2, len(failed_rows) + 1), len(failed_detail_headers), 'FailedTestsTable')
    autosize(failed_sheet, max_width=72)
    for row in failed_sheet.iter_rows(min_row=2):
        if row[3].value == 'FAIL':
            row[3].fill = PatternFill('solid', fgColor=PALETTE['fail_bg'])
            row[3].font = Font(name='Aptos', bold=True, color=PALETTE['fail'])
        for cell in row:
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    for ws in (dashboard, all_tests, failed_sheet):
        ws.sheet_view.zoomScale = 90

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    args = parse_args()
    summary = load_summary(Path(args.combined_json))
    build_workbook(summary, Path(args.output), args.title, args.subtitle)
    print(f'OUTPUT_XLSX={args.output}')


if __name__ == '__main__':
    main()
