#!/usr/bin/env python3
"""Build the Hydro-QA workflow-evidence Excel from the evidence-gather workflow output.

Usage: python build_evidence_excel.py <workflow_output.json> <out.xlsx>
Produces: Summary dashboard + Screenshots proof sheet (embedded validated thumbnails) +
per-workflow detail sheets + Automation (non-test) sheet. Verdicts/screenshot-validity are
visually proven, not asserted: each embedded screenshot was rendered + judged by an agent.
"""
import json, os, re, sys, tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

GREEN = PatternFill("solid", fgColor="C6EFCE"); GREEN_F = Font(color="006100", bold=True)
AMBER = PatternFill("solid", fgColor="FFEB9C"); AMBER_F = Font(color="9C6500", bold=True)
RED = PatternFill("solid", fgColor="FFC7CE"); RED_F = Font(color="9C0006", bold=True)
HEAD = PatternFill("solid", fgColor="1F4E78"); HEAD_F = Font(color="FFFFFF", bold=True, size=11)
TITLE_F = Font(bold=True, size=15, color="1F4E78")
WRAP = Alignment(wrap_text=True, vertical="top")
THIN = Border(*[Side(style="thin", color="D9D9D9")] * 4)

VERDICT_STYLE = {"Proven": (GREEN, GREEN_F), "Weak": (AMBER, AMBER_F), "Broken": (RED, RED_F),
                 "strong": (GREEN, GREEN_F), "weak": (AMBER, AMBER_F), "none": (RED, RED_F), "unknown": (AMBER, AMBER_F)}
SS_STYLE = {"correct": (GREEN, GREEN_F), "na": (AMBER, AMBER_F)}  # anything else = RED


def safe_sheet(name):
    m = re.search(r"([\w-]+)\.yml", name)
    base = m.group(1) if m else re.sub(r"[\\/?*\[\]:]", "", name.split("—")[0].split("(")[0].strip())
    return (base or "wf")[:31]


def short_wf(name):
    m = re.search(r"([\w-]+\.yml)", name)
    return m.group(1) if m else name.split("—")[0].strip()[:40]


def style_cell(c, fill=None, font=None, wrap=True):
    if fill: c.fill = fill
    if font: c.font = font
    if wrap: c.alignment = WRAP
    c.border = THIN


def build(out_json, out_xlsx):
    d = json.load(open(out_json, encoding="utf-8"))
    res = d.get("result", d)
    te = res.get("testEvidence", [])
    pop = res.get("populator", [])
    wb = Workbook()
    tmpdir = tempfile.mkdtemp(prefix="evid_thumbs_")

    # ---------- Summary ----------
    ws = wb.active; ws.title = "Summary"
    ws["A1"] = "Hydro-QA Workflow Evidence — proof each test actually works"; ws["A1"].font = TITLE_F
    ws["A2"] = ("Generated 2026-05-28. Each verdict is backed by real artifacts; every embedded screenshot was RENDERED and judged "
                "(correct screen vs blank/ANR/login-error/stale/wrong-screen) — see the 'Screenshots' sheet. Assertion strength = does the "
                "workflow actually FAIL on a regression, or report green regardless (fail-open).")
    ws["A2"].font = Font(italic=True, color="595959"); ws.merge_cells("A2:H2"); ws["A2"].alignment = WRAP
    proven = sum(1 for e in te if e["verdict"] == "Proven")
    ssall = [s for e in te for s in e.get("screenshots", [])]
    ssok = sum(1 for s in ssall if s["verdict"] in ("correct", "na"))
    ws["A4"] = (f"HEADLINE: {proven}/{len(te)} test workflows are PROVEN; {len(te)-proven} are WEAK (assertion/evidence gaps). "
                f"Screenshots validated: {ssok}/{len(ssall)} show the correct screen.")
    ws["A4"].font = Font(bold=True, size=12, color="9C0006" if proven < len(te) else "006100"); ws.merge_cells("A4:H4"); ws["A4"].alignment = WRAP
    ws["A5"] = ("LEGEND  —  Verdict: PROVEN = real fail-closed assertions + validated evidence + stable.  "
                "WEAK = the workflow RUNS (not broken) but its assertions are weak/absent (can report green even through a regression) "
                "or its evidence isn't validated.  BROKEN = fails / stale / fakes green.   "
                "Assertions: strong/weak/none = does the job actually go RED on a regression?   "
                "Screenshots: validated = an agent rendered it and confirmed the correct screen (not blank/ANR/login-error/stale/wrong-screen).")
    ws["A5"].font = Font(italic=True, size=9, color="595959"); ws.merge_cells("A5:H5"); ws["A5"].alignment = WRAP; ws.row_dimensions[5].height = 56

    hdr = ["Workflow", "Verdict", "Assertions", "CI stability (recent)", "Screenshots (valid/total)", "What it tests", "Key gaps", "Detail sheet"]
    r = 7
    for i, h in enumerate(hdr, 1):
        c = ws.cell(r, i, h); style_cell(c, HEAD, HEAD_F)
    for e in te:
        r += 1
        ss = e.get("screenshots", [])
        ssg = sum(1 for s in ss if s["verdict"] in ("correct", "na"))
        cells = [
            (short_wf(e["workflow"]), None, None),
            (e["verdict"], *VERDICT_STYLE.get(e["verdict"], (None, None))),
            (e["assertionVerdict"], *VERDICT_STYLE.get(e["assertionVerdict"], (None, None))),
            (e.get("flakeSummary", "")[:300], None, None),
            (f"{ssg}/{len(ss)}" if ss else "n/a (report-based)", *( (GREEN, GREEN_F) if ss and ssg == len(ss) else (AMBER, AMBER_F) if ss else (None, None))),
            (e.get("whatItTests", "")[:400], None, None),
            (" • ".join(e.get("gaps", []))[:400], None, None),
            (safe_sheet(e["workflow"]), None, None),
        ]
        for i, (val, fill, font) in enumerate(cells, 1):
            style_cell(ws.cell(r, i, val), fill, font)
        ws.row_dimensions[r].height = 90
    widths = [26, 10, 11, 46, 16, 50, 46, 18]
    for i, w in enumerate(widths, 1): ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A8"

    # ---------- Screenshots (proof) ----------
    sh = wb.create_sheet("Screenshots")
    sh["A1"] = "Screenshot validation — each was rendered and judged (the proof-of-proof)"; sh["A1"].font = TITLE_F
    sh.merge_cells("A1:E1")
    shh = ["Workflow", "Screenshot", "Verdict", "What it actually shows", "Thumbnail"]
    for i, h in enumerate(shh, 1): style_cell(sh.cell(3, i, h), HEAD, HEAD_F)
    sh.column_dimensions["A"].width = 24; sh.column_dimensions["B"].width = 34
    sh.column_dimensions["C"].width = 14; sh.column_dimensions["D"].width = 52; sh.column_dimensions["E"].width = 42
    rr = 4
    for e in te:
        for s in e.get("screenshots", []):
            fill, font = SS_STYLE.get(s["verdict"], (RED, RED_F))
            style_cell(sh.cell(rr, 1, short_wf(e["workflow"])), None, None)
            style_cell(sh.cell(rr, 2, s["file"]), None, None)
            style_cell(sh.cell(rr, 3, s["verdict"]), fill, font)
            style_cell(sh.cell(rr, 4, s["reason"]), None, None)
            lp = s.get("localPath", "")
            sh.row_dimensions[rr].height = 200
            try:
                if lp and os.path.exists(lp):
                    im = PILImage.open(lp); im.thumbnail((260, 260))
                    tp = os.path.join(tmpdir, f"t{rr}.png"); im.save(tp)
                    xi = XLImage(tp); sh.add_image(xi, f"E{rr}")
                else:
                    style_cell(sh.cell(rr, 5, "(image missing)"), None, None)
            except Exception as ex:
                style_cell(sh.cell(rr, 5, f"(embed failed: {ex})"), None, None)
            rr += 1
    sh.freeze_panes = "A4"

    # ---------- per-workflow detail ----------
    for e in te:
        ds = wb.create_sheet(safe_sheet(e["workflow"]))
        ds["A1"] = short_wf(e["workflow"]); ds["A1"].font = TITLE_F; ds.merge_cells("A1:B1")
        ds.column_dimensions["A"].width = 22; ds.column_dimensions["B"].width = 110
        rows = [
            ("Verdict", e["verdict"]), ("Assertion strength", e["assertionVerdict"]),
            ("Purpose", e.get("purpose", "")), ("What it tests", e.get("whatItTests", "")),
            ("Assertions (fail-closed?)", e.get("assertionStrength", "")),
            ("CI stability", e.get("flakeSummary", "")), ("Coverage", e.get("coverage", "")),
            ("Rationale", e.get("rationale", "")), ("Gaps", " • ".join(e.get("gaps", []))),
            ("Artifacts", "\n".join(e.get("artifacts", []))),
            ("CI history", "\n".join(f"{h['runId']}  {h['conclusion']}  {h['date']}" for h in e.get("ciHistory", []))),
        ]
        ri = 3
        for k, v in rows:
            kc = ds.cell(ri, 1, k); kc.font = Font(bold=True); kc.alignment = WRAP; kc.fill = PatternFill("solid", fgColor="EAF1FB"); kc.border = THIN
            vcell = ds.cell(ri, 2, str(v)); style_cell(vcell)
            if k == "Verdict": vcell.fill, vcell.font = VERDICT_STYLE.get(v, (None, None))
            ds.row_dimensions[ri].height = max(30, min(360, 16 + len(str(v)) // 110 * 15))
            ri += 1

    # ---------- Automation (non-test) ----------
    az = wb.create_sheet("Automation (non-test)")
    az["A1"] = "claude-populator* — DATA-PATCHING automation, NOT QA tests (classified, not evidence-graded)"; az["A1"].font = TITLE_F
    az.merge_cells("A1:D1")
    for i, h in enumerate(["Workflow", "Purpose", "Classification", "Last run"], 1): style_cell(az.cell(3, i, h), HEAD, HEAD_F)
    for w, wd in zip([28, 60, 50, 26], "ABCD"): az.column_dimensions[wd].width = w
    ar = 4
    for p in pop:
        for i, key in enumerate(["workflow", "purpose", "classification", "lastRun"], 1):
            style_cell(az.cell(ar, i, str(p.get(key, ""))))
        az.row_dimensions[ar].height = 60; ar += 1

    wb.save(out_xlsx)
    print(f"WROTE {out_xlsx}  ({len(te)} test sheets + Screenshots[{len(ssall)}] + Automation[{len(pop)}])")


if __name__ == "__main__":
    build(sys.argv[1], sys.argv[2])
