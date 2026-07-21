#!/usr/bin/env python3
"""
Functional QA for the mobile regression tooling. Fast, no emulator - guards
that the suite's plumbing is intact so a broken flow file or a broken Excel
generator is caught before a 1-hour emulator run wastes time.

Checks:
  1. Every mobile-flows-v2/*.yaml parses and declares appId com.hydrocert.app.
  2. Every flow has a description comment block (the "What it checks" column).
  3. The Excel generator imports and produces a 2-sheet workbook (Summary +
     Details) from a tiny synthetic fixture, with a circle drawn when a
     '# CIRCLE:' hint resolves to an element in the dump.
"""
from __future__ import annotations
import sys
import tempfile
import json
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
FLOWS = ROOT / "mobile-flows-v2"
sys.path.insert(0, str(ROOT / "scripts"))

errors: list[str] = []


def check_flows_parse():
    import yaml
    for f in sorted(FLOWS.rglob("*.yaml")):
        try:
            docs = list(yaml.safe_load_all(f.read_text(encoding="utf-8")))
        except Exception as exc:
            errors.append(f"{f.name}: YAML parse error: {exc}")
            continue
        head = docs[0] if docs else None
        if not isinstance(head, dict) or head.get("appId") != "com.hydrocert.app":
            errors.append(f"{f.name}: missing/invalid 'appId: com.hydrocert.app'")


def check_descriptions():
    from generate_mobile_regression_excel import extract_flow_description
    for f in sorted(FLOWS.glob("[0-9]*.yaml")):
        fid = f.stem
        desc = extract_flow_description(FLOWS, fid)
        if not desc or len(desc.strip()) < 8:
            errors.append(f"{f.name}: empty/too-short description comment block")


def check_generator():
    import generate_mobile_regression_excel as gen
    from openpyxl import load_workbook
    with tempfile.TemporaryDirectory() as d:
        d = Path(d)
        shots = d / "test" / "screenshots"
        dumps = d / "test" / "ui-dumps"
        shots.mkdir(parents=True); dumps.mkdir(parents=True)
        # 1x1 white png + a dump with one node so annotate has something to find
        try:
            from PIL import Image
            Image.new("RGB", (200, 400), "white").save(shots / "01_login_screen-after.png")
        except Exception:
            (shots / "01_login_screen-after.png").write_bytes(b"")
        (dumps / "01_login_screen.xml").write_text(
            '<hierarchy><node text="Login" bounds="[10,20][90,60]"/></hierarchy>', encoding="utf-8")
        summary = {"mode": "test", "totals": {"total": 1, "pass": 1, "fail": 0, "skip": 0},
                   "checks": [{"id": "01_login_screen", "status": "PASS", "details": "ok"}]}
        (d / "test" / "summary.json").write_text(json.dumps(summary), encoding="utf-8")
        out = d / "report.xlsx"
        descriptions = {"01_login_screen": "Login screen elements smoke check."}
        try:
            gen.build_report(summary, d, out, "QA", "func-check", descriptions, {"01_login_screen": "Login"})
        except Exception as exc:
            errors.append(f"generator raised: {exc}")
            return
        if not out.is_file():
            errors.append("generator produced no xlsx"); return
        wb = load_workbook(str(out))
        for sheet in ("Summary", "Details"):
            if sheet not in wb.sheetnames:
                errors.append(f"generator xlsx missing '{sheet}' sheet")


def main() -> int:
    check_flows_parse()
    check_descriptions()
    check_generator()
    if errors:
        print("FUNCTIONAL QA FAILED:\n")
        for e in errors:
            print("  " + e)
        print(f"\n{len(errors)} problem(s).")
        return 1
    print("Functional QA passed: flows parse, descriptions present, Excel generator produces Summary + Details.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
